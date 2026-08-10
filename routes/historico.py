"""Rotas do módulo historico."""

from flask import Blueprint

from backend import *  # noqa: F401,F403
from decorators import login_required, permission_required


bp = Blueprint("historico", __name__)


@bp.before_request
@login_required
@permission_required("historico")
def enforce_history_permissions():
    return None


@bp.route("/historico")
def historico():
    dados = obter_historico()
    agora = datetime.now()
    hoje_total = sum(
        1 for item in dados
        if item["execucao_dt"] and item["execucao_dt"].date() == agora.date()
    )
    mes_total = sum(
        1 for item in dados
        if item["execucao_dt"]
        and item["execucao_dt"].year == agora.year
        and item["execucao_dt"].month == agora.month
    )
    ultimo = dados[0] if dados else None

    return render_template(
        "historico.html",
        dados=dados,
        total=len(dados),
        hoje_total=hoje_total,
        mes_total=mes_total,
        ultimo=ultimo,
        timeline=dados[:5],
        busca=request.args.get("busca", ""),
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y"),
        data_hoje_dt=agora.date()
    )


@bp.route("/historico/exportar/excel")
def exportar_historico_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    dados = obter_historico()
    preferencias = carregar_configuracoes()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Histórico"
    sheet.append(["Colaborador", "Data do bloqueio", "Executado em", "Status"])

    for item in dados:
        sheet.append([
            item["nome"],
            item["data_bloqueio"],
            item["data_execucao"],
            item["status"]
        ])

    header_fill = PatternFill("solid", fgColor="082F5B")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")

    if preferencias.get("excel_congelar_cabecalho") == "1":
        sheet.freeze_panes = "A2"
    if preferencias.get("excel_autofiltro") == "1":
        sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 26
    if preferencias.get("excel_ajustar_colunas") == "1":
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 23
        sheet.column_dimensions["D"].width = 16

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    arquivo = BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)
    registrar_auditoria(
        "Exportou Excel",
        f"Histórico de bloqueios · {len(dados)} registro(s)"
    )

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"historico_bloqueios_{datetime.now():%Y%m%d}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@bp.route("/historico/exportar/pdf")
def exportar_historico_pdf():
    registrar_auditoria(
        "Gerou PDF",
        f"Histórico de bloqueios · {len(obter_historico())} registro(s)"
    )
    arquivo = BytesIO(gerar_pdf_historico(obter_historico()))
    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"historico_bloqueios_{datetime.now():%Y%m%d}.pdf",
        mimetype="application/pdf"
    )
