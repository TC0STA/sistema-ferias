"""Rotas do módulo relatorios."""

from flask import Blueprint

from backend import *  # noqa: F401,F403


bp = Blueprint("relatorios", __name__)


@bp.route("/relatorios")
def relatorios():
    dados = obter_dados_relatorios()
    return render_template(
        "relatorios.html",
        **dados,
        data_atual=dados["agora"].strftime("%d/%m/%Y %H:%M"),
        data_hoje=dados["agora"].strftime("%d/%m/%Y")
    )


@bp.route("/relatorios/exportar/csv")
def exportar_relatorios_csv():
    dados = obter_dados_relatorios()
    texto = StringIO()
    texto.write("\ufeff")
    writer = csv.writer(texto, delimiter=";")
    writer.writerow([
        "Colaborador", "Departamento", "Início", "Fim",
        "Bloqueio", "Dias", "Status"
    ])
    writer.writerows(linhas_exportacao_relatorios(dados))
    arquivo = BytesIO(texto.getvalue().encode("utf-8"))
    registrar_auditoria(
        "Exportou CSV",
        f"Relatório gerencial · {dados['periodo_nome']} · {len(dados['registros'])} registro(s)"
    )
    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"relatorio_ferias_{dados['periodo'].replace('-', '_')}.csv",
        mimetype="text/csv; charset=utf-8"
    )


@bp.route("/relatorios/exportar/excel")
def exportar_relatorios_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    dados = obter_dados_relatorios()
    preferencias = carregar_configuracoes()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatório de férias"
    sheet.append([
        "Colaborador", "Departamento", "Início", "Fim",
        "Bloqueio", "Dias", "Status"
    ])
    for linha in linhas_exportacao_relatorios(dados):
        sheet.append(linha)

    header_fill = PatternFill("solid", fgColor="082F5B")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    if preferencias.get("excel_congelar_cabecalho") == "1":
        sheet.freeze_panes = "A2"
    if preferencias.get("excel_autofiltro") == "1":
        sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 26
    if preferencias.get("excel_ajustar_colunas") == "1":
        for coluna, largura in {
            "A": 34, "B": 24, "C": 15, "D": 15,
            "E": 15, "F": 10, "G": 17
        }.items():
            sheet.column_dimensions[coluna].width = largura

    arquivo = BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)
    registrar_auditoria(
        "Exportou Excel",
        f"Relatório gerencial · {dados['periodo_nome']} · {len(dados['registros'])} registro(s)"
    )
    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"relatorio_ferias_{dados['periodo'].replace('-', '_')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@bp.route("/relatorios/exportar/pdf")
def exportar_relatorios_pdf():
    dados = obter_dados_relatorios()
    registrar_auditoria(
        "Gerou PDF",
        f"Relatório gerencial · {dados['periodo_nome']} · {len(dados['registros'])} registro(s)"
    )
    return send_file(
        BytesIO(gerar_pdf_relatorios(dados)),
        as_attachment=True,
        download_name=f"relatorio_ferias_{dados['periodo'].replace('-', '_')}.pdf",
        mimetype="application/pdf"
    )
