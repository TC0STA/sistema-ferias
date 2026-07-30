from flask import Flask, render_template, request, send_from_directory, send_file, redirect, url_for
import pandas as pd
import os
import sqlite3
import unicodedata
import calendar as calendar_module
import csv
import json
import zipfile
import shutil
import time
import uuid
from io import BytesIO
from io import StringIO
from datetime import datetime, timedelta
from urllib.parse import quote

from services.events import EventBus
from services.audit_service import AuditService
from services.column_alias_store import ColumnAliasStore
from services.import_service import ImportService
from services.import_plugin import PluginExecutionError

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
DATABASE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "database", "ferias.db")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
VALIDACOES_DIR = os.path.join(BASE_DIR, "uploads", ".validacoes")
BACKUPS_DIR = os.path.join(BASE_DIR, "backups")
IMPORT_EVENTS = EventBus()

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(os.path.dirname(DATABASE_PATH), exist_ok=True)

CONFIGURACOES_PADRAO = {
    "nome_empresa": "Grupo Fokus Logística",
    "nome_sistema": "Fokus Férias",
    "versao": "2.0",
    "banco": "SQLite",
    "ambiente": "Produção",
    "notificacoes_ativas": "1",
    "hora_rotina": "08:00",
    "dias_antes_bloqueio": "3",
    "som_notificacoes": "1",
    "pasta_padrao": "uploads",
    "substituir_planilha": "0",
    "validar_planilha": "1",
    "mostrar_resumo_importacao": "1",
    "backup_antes_importacao": "1",
    "mostrar_popup": "1",
    "notificar_importacao": "1",
    "tema": "claro",
    "cor_principal": "azul-fokus",
    "tamanho_fonte": "normal",
    "pdf_logo": "1",
    "pdf_rodape": "1",
    "pdf_numero_pagina": "1",
    "pdf_data": "1",
    "excel_autofiltro": "1",
    "excel_ajustar_colunas": "1",
    "excel_congelar_cabecalho": "1",
    "limpeza_backups": "1",
    "manter_backups": "10"
}


def inicializar_tabelas_sistema():
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS auditoria (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data_hora TEXT NOT NULL,
            acao TEXT NOT NULL,
            detalhe TEXT,
            usuario TEXT NOT NULL,
            ip TEXT
        )
    """)
    colunas_auditoria = {
        item[1] for item in cursor.execute("PRAGMA table_info(auditoria)").fetchall()
    }
    if "ip" not in colunas_auditoria:
        cursor.execute("ALTER TABLE auditoria ADD COLUMN ip TEXT")
    if "resultado" not in colunas_auditoria:
        cursor.execute(
            "ALTER TABLE auditoria ADD COLUMN resultado TEXT NOT NULL DEFAULT 'Sucesso'"
        )
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS configuracoes (
            chave TEXT PRIMARY KEY,
            valor TEXT NOT NULL,
            atualizado_em TEXT NOT NULL
        )
    """)
    agora = datetime.now().isoformat()
    for chave, valor in CONFIGURACOES_PADRAO.items():
        cursor.execute(
            "INSERT OR IGNORE INTO configuracoes (chave, valor, atualizado_em) VALUES (?, ?, ?)",
            (chave, valor, agora)
        )
    conn.commit()
    conn.close()
    ColumnAliasStore(DATABASE_PATH).ensure_schema()


def registrar_auditoria(
    acao, detalhe="", usuario="Tiago Costa", ip=None, resultado="Sucesso"
):
    inicializar_tabelas_sistema()
    if ip is None:
        try:
            ip = request.headers.get("X-Forwarded-For", request.remote_addr or "Local")
            ip = ip.split(",")[0].strip()
        except RuntimeError:
            ip = "Local"
    AuditService(DATABASE_PATH).record(
        acao,
        detalhe,
        user=usuario,
        ip=ip,
        result=resultado
    )


def obter_ip_requisicao():
    try:
        ip = request.headers.get("X-Forwarded-For", request.remote_addr or "Local")
        return ip.split(",")[0].strip()
    except RuntimeError:
        return "Local"


def obter_motor_importacao(*, readonly=False):
    return ImportService(
        database_path=DATABASE_PATH,
        backups_root=BACKUPS_DIR,
        event_bus=IMPORT_EVENTS,
        persist_plugin_registry=not readonly
    )


def auditar_importacao_concluida(evento):
    for nome in evento.get("bloqueios_registrados", []):
        registrar_auditoria(
            "Executou bloqueio",
            f"Colaborador: {nome}",
            usuario=evento["usuario"],
            ip=evento["ip"]
        )


IMPORT_EVENTS.subscribe(
    "importacao_concluida",
    auditar_importacao_concluida
)


def carregar_configuracoes():
    inicializar_tabelas_sistema()
    conn = sqlite3.connect(DATABASE_PATH)
    valores = dict(conn.execute("SELECT chave, valor FROM configuracoes").fetchall())
    conn.close()
    return {**CONFIGURACOES_PADRAO, **valores}


def obter_pasta_planilhas():
    pasta = carregar_configuracoes().get("pasta_padrao", UPLOAD_FOLDER).strip()
    caminho = pasta if os.path.isabs(pasta) else os.path.join(BASE_DIR, pasta)
    os.makedirs(caminho, exist_ok=True)
    return caminho


def listar_backups():
    pasta = BACKUPS_DIR
    if not os.path.isdir(pasta):
        return []
    arquivos = []
    for raiz, _, nomes in os.walk(pasta):
        arquivos.extend(
            os.path.join(raiz, nome)
            for nome in nomes
            if nome.lower().endswith((".zip", ".db"))
        )
    return sorted(arquivos, key=os.path.getmtime, reverse=True)


def criar_backup(origem="manual"):
    os.makedirs(BACKUPS_DIR, exist_ok=True)
    nome_arquivo = f"backup_ferias_{datetime.now():%Y%m%d_%H%M%S_%f}.zip"
    caminho_backup = os.path.join(BACKUPS_DIR, nome_arquivo)
    configuracoes_atuais = carregar_configuracoes()

    with zipfile.ZipFile(caminho_backup, "w", zipfile.ZIP_DEFLATED) as arquivo:
        if os.path.exists(DATABASE_PATH):
            arquivo.write(DATABASE_PATH, "database/ferias.db")
        arquivo.writestr(
            "configuracoes/configuracoes.json",
            json.dumps(configuracoes_atuais, ensure_ascii=False, indent=2)
        )
        pasta_logs = os.path.join(BASE_DIR, "logs")
        if os.path.isdir(pasta_logs):
            for raiz, _, nomes in os.walk(pasta_logs):
                for nome in nomes:
                    caminho_log = os.path.join(raiz, nome)
                    relativo = os.path.relpath(caminho_log, pasta_logs)
                    arquivo.write(caminho_log, os.path.join("logs", relativo))

    if configuracoes_atuais.get("limpeza_backups") == "1":
        try:
            manter = max(1, int(configuracoes_atuais.get("manter_backups", "10")))
        except ValueError:
            manter = 10
        for antigo in listar_backups()[manter:]:
            os.remove(antigo)

    registrar_auditoria(
        "Backup",
        f"Backup {origem}: {nome_arquivo}",
        usuario="Sistema" if origem == "automático" else "Tiago Costa"
    )
    return caminho_backup


def limpar_backups_antigos():
    configuracoes = carregar_configuracoes()
    try:
        manter = max(1, int(configuracoes.get("manter_backups", "10")))
    except ValueError:
        manter = 10
    removidos = 0
    for caminho in listar_backups()[manter:]:
        os.remove(caminho)
        removidos += 1
    return removidos


def obter_saude_sistema():
    pasta = obter_pasta_planilhas()
    caminho_planilha = planilha_mais_recente()
    backups = listar_backups()
    agora = datetime.now()
    itens = []

    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.execute("SELECT 1").fetchone()
        conn.close()
        itens.append(("Banco SQLite", "Funcionando", "ok", "database"))
    except sqlite3.Error:
        itens.append(("Banco SQLite", "Falha na conexão", "error", "database"))

    if os.path.isdir(pasta):
        itens.append(("Pasta uploads", "OK", "ok", "folder-check"))
    else:
        itens.append(("Pasta uploads", "Não encontrada", "error", "folder-x"))

    permissao = os.access(pasta, os.R_OK | os.W_OK)
    itens.append((
        "Permissões", "OK" if permissao else "Sem acesso de escrita",
        "ok" if permissao else "error", "shield-check"
    ))

    if caminho_planilha:
        momento = datetime.fromtimestamp(os.path.getmtime(caminho_planilha))
        dias = (agora - momento).days
        nivel = "ok" if dias <= 7 else "warning"
        texto = "Sucesso" if dias <= 7 else "Nenhuma planilha nesta semana"
        itens.append(("Última importação", texto, nivel, "file-check-2"))
    else:
        itens.append((
            "Última importação", "Nenhuma planilha importada",
            "warning", "file-warning"
        ))

    if backups:
        momento = datetime.fromtimestamp(os.path.getmtime(backups[0]))
        dias = (agora - momento).days
        nivel = "ok" if dias <= 30 else "error"
        texto = "Atualizado" if dias <= 30 else "Há mais de 30 dias"
        itens.append(("Backup", texto, nivel, "database-backup"))
    else:
        itens.append(("Backup", "Ainda não realizado", "error", "database-zap"))

    livres = shutil.disk_usage(BASE_DIR).free / (1024 ** 3)
    espaco_texto = f"{livres:.1f}".replace(".", ",")
    itens.append(("Espaço em disco", f"{espaco_texto} GB livres", "ok", "hard-drive"))
    return [
        {"nome": nome, "status": status, "nivel": nivel, "icone": icone}
        for nome, status, nivel, icone in itens
    ]

def normalizar_texto(valor):
    texto = str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(
        caractere
        for caractere in texto
        if not unicodedata.combining(caractere)
    )

def contem_palavra(texto, palavras):
    return any(palavra in texto for palavra in palavras)

def encontrar_coluna(df, opcoes):
    for coluna in df.columns:
        coluna_normalizada = normalizar_texto(coluna)

        if coluna_normalizada in opcoes:
            return coluna

    for coluna in df.columns:
        coluna_normalizada = normalizar_texto(coluna)

        if contem_palavra(coluna_normalizada, opcoes):
            return coluna

    return None

def _read_excel_sheet(caminho, sheet_name):
    bruto = pd.read_excel(caminho, sheet_name=sheet_name, header=None)

    for indice, row in bruto.iterrows():
        valores = [
            normalizar_texto(valor)
            for valor in row.tolist()
            if str(valor).strip().lower() != "nan"
        ]

        tem_nome = any(
            valor in [
                "nome", "funcionario", "colaborador", "usuario",
                "nome do colaborador", "nome do funcionario"
            ]
            for valor in valores
        )
        tem_inicio = any(
            valor in [
                "data inicio", "inicio", "inicio das ferias",
                "inicio ferias", "ferias inicio", "data inicial",
                "data de inicio"
            ] or "inicio" in valor
            for valor in valores
        )
        tem_fim = any(
            valor in [
                "data fim", "fim", "retorno", "data retorno", "volta",
                "fim das ferias", "fim ferias", "data final", "data de fim"
            ] or "fim" in valor
            for valor in valores
        )

        if tem_nome and tem_inicio and tem_fim:
            df = pd.read_excel(caminho, sheet_name=sheet_name, header=indice)
            df.columns = df.columns.str.strip()
            return df

    df = pd.read_excel(caminho, sheet_name=sheet_name)
    df.columns = df.columns.str.strip()
    return df


def ler_excel_com_cabecalho(caminho):
    excel = pd.ExcelFile(caminho)
    dataframes = []

    for sheet_name in excel.sheet_names:
        df = _read_excel_sheet(caminho, sheet_name)
        if not df.empty:
            dataframes.append(df)

    if not dataframes:
        df = pd.read_excel(caminho)
        df.columns = df.columns.str.strip()
        return df

    if len(dataframes) == 1:
        return dataframes[0]

    return pd.concat(dataframes, ignore_index=True, sort=False)

def carregar_planilha(caminho, mapeamento=None):
    df = ler_excel_com_cabecalho(caminho)

    mapeamento = mapeamento or {}
    coluna_nome = mapeamento.get("nome") or encontrar_coluna(
        df,
        [
            "nome", "funcionario", "colaborador", "usuario",
            "nome do colaborador", "nome do funcionario"
        ]
    )
    coluna_inicio = mapeamento.get("inicio") or encontrar_coluna(
        df,
        [
            "inicio", "inicio ferias", "inicio das ferias", "data inicio",
            "data de inicio", "ferias inicio", "data inicial"
        ]
    )
    coluna_fim = mapeamento.get("fim") or encontrar_coluna(
        df,
        [
            "fim", "fim ferias", "fim das ferias", "data fim",
            "data de fim", "retorno", "data retorno", "volta", "data final"
        ]
    )
    coluna_bloqueio = encontrar_coluna(
        df,
        ["data de bloqueio", "bloqueio", "data bloqueio"]
    )

    if coluna_nome is None or coluna_inicio is None or coluna_fim is None:
        raise ValueError(
            "A planilha precisa ter colunas de Nome, Inicio e Fim."
        )

    colunas = {
        coluna_nome: "Nome",
        coluna_inicio: "Inicio",
        coluna_fim: "Fim"
    }

    if coluna_bloqueio is not None:
        colunas[coluna_bloqueio] = "Data de Bloqueio"

    for origem, destino in colunas.items():
        if (
            origem != destino
            and destino in df.columns
            and destino not in colunas
        ):
            df = df.drop(columns=[destino])
    df = df.rename(columns=colunas)

    df["Nome"] = df["Nome"].astype(str).str.strip()
    df = df[
        (df["Nome"] != "")
        & (df["Nome"].str.lower() != "nan")
    ].copy()

    df["Inicio"] = pd.to_datetime(
        df["Inicio"],
        errors="coerce",
        dayfirst=True
    )

    df["Fim"] = pd.to_datetime(
        df["Fim"],
        errors="coerce",
        dayfirst=True
    )

    if "Data de Bloqueio" not in df.columns:
        dias_antes = int(
            carregar_configuracoes().get("dias_antes_bloqueio", "1")
        )
        df["Data de Bloqueio"] = df["Inicio"] - timedelta(days=dias_antes)

    df["Data de Bloqueio"] = pd.to_datetime(
        df["Data de Bloqueio"],
        errors="coerce",
        dayfirst=True
    )

    return df

def planilhas_importadas():
    pasta_planilhas = obter_pasta_planilhas()
    arquivos = [
        os.path.join(pasta_planilhas, nome)
        for nome in os.listdir(pasta_planilhas)
        if nome.lower().endswith((".xlsx", ".xls"))
    ]
    return sorted(arquivos, key=os.path.getmtime)


def planilha_mais_recente():
    arquivos = planilhas_importadas()
    if not arquivos:
        return None
    return arquivos[-1]


def carregar_planilhas(caminhos, mapeamento=None):
    frames = []
    for caminho in caminhos:
        try:
            df = carregar_planilha(caminho, mapeamento=mapeamento)
        except Exception:
            continue
        if not df.empty:
            frames.append(df)

    if not frames:
        return pd.DataFrame(columns=["Nome", "Inicio", "Fim", "Data de Bloqueio"])

    df = pd.concat(frames, ignore_index=True, sort=False)
    if {"Nome", "Inicio", "Fim", "Data de Bloqueio"}.issubset(df.columns):
        df = df.drop_duplicates(subset=["Nome", "Inicio", "Fim", "Data de Bloqueio"])
    return df.reset_index(drop=True)


def limpar_validacoes_temporarias(limite_segundos=3600):
    os.makedirs(VALIDACOES_DIR, exist_ok=True)
    agora = time.time()
    for nome in os.listdir(VALIDACOES_DIR):
        caminho = os.path.join(VALIDACOES_DIR, nome)
        if (
            os.path.isfile(caminho)
            and agora - os.path.getmtime(caminho) > limite_segundos
        ):
            os.remove(caminho)


def token_validacao_seguro(token):
    return (
        isinstance(token, str)
        and len(token) == 32
        and all(caractere in "0123456789abcdef" for caractere in token)
    )


def caminho_metadados_validacao(token):
    return os.path.join(VALIDACOES_DIR, f"{token}.json")


def carregar_validacao_temporaria(token):
    if not token_validacao_seguro(token):
        return None
    caminho_metadados = caminho_metadados_validacao(token)
    if not os.path.isfile(caminho_metadados):
        return None
    try:
        with open(caminho_metadados, "r", encoding="utf-8") as arquivo:
            metadados = json.load(arquivo)
    except (OSError, json.JSONDecodeError):
        return None
    extensao = metadados.get("extensao", "")
    caminho = os.path.join(VALIDACOES_DIR, f"{token}{extensao}")
    if not os.path.isfile(caminho):
        return None
    metadados["caminho"] = caminho
    metadados["caminho_metadados"] = caminho_metadados
    return metadados


def formatar_datas(df):
    df = df.copy()

    for coluna in ["Inicio", "Fim", "Data de Bloqueio"]:
        if coluna in df.columns:
            df[coluna] = df[coluna].dt.strftime("%d/%m/%Y")
            df[coluna] = df[coluna].fillna("")

    return df


def obter_dados_relatorios():
    agora = datetime.now()
    hoje = agora.date()
    periodo = request.args.get("periodo", agora.strftime("%Y-%m"))
    departamento_filtro = request.args.get("departamento", "").strip()
    status_filtro = request.args.get("status", "").strip()
    pesquisa = request.args.get("q", "").strip()

    try:
        ano_selecionado, mes_selecionado = [int(parte) for parte in periodo.split("-", 1)]
        primeiro_dia = datetime(ano_selecionado, mes_selecionado, 1).date()
    except (TypeError, ValueError):
        ano_selecionado, mes_selecionado = hoje.year, hoje.month
        periodo = agora.strftime("%Y-%m")
        primeiro_dia = hoje.replace(day=1)

    ultimo_dia = (
        datetime(ano_selecionado + 1, 1, 1).date()
        if mes_selecionado == 12
        else datetime(ano_selecionado, mes_selecionado + 1, 1).date()
    ) - timedelta(days=1)

    colaboradores = obter_colaboradores()
    departamentos = sorted({
        item["departamento"] for item in colaboradores
        if item["departamento"] != "Não informado"
    })

    if departamento_filtro:
        colaboradores = [
            item for item in colaboradores
            if item["departamento"] == departamento_filtro
        ]
    if pesquisa:
        pesquisa_normalizada = normalizar_texto(pesquisa)
        colaboradores = [
            item for item in colaboradores
            if pesquisa_normalizada in normalizar_texto(item["nome"])
        ]
    if status_filtro:
        colaboradores = [
            item for item in colaboradores
            if item["status_classe"] == status_filtro
        ]

    registros = []
    for colaborador in colaboradores:
        for periodo_ferias in colaborador["periodos"]:
            inicio = periodo_ferias["inicio"]
            fim = periodo_ferias["fim"]
            status_classe = (
                "active" if inicio <= hoje <= fim
                else "scheduled" if inicio > hoje
                else "completed"
            )
            status = {
                "active": "Em férias",
                "scheduled": "Programada",
                "completed": "Concluída"
            }[status_classe]
            registros.append({
                "nome": colaborador["nome"],
                "departamento": colaborador["departamento"],
                "inicio": inicio,
                "fim": fim,
                "bloqueio": periodo_ferias["bloqueio"],
                "status": status,
                "status_classe": status_classe,
                "dias": (fim - inicio).days + 1
            })

    registros_periodo = [
        item for item in registros
        if item["inicio"] <= ultimo_dia and item["fim"] >= primeiro_dia
    ]

    nomes_filtrados = {item["nome"] for item in colaboradores}
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bloqueios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT,
            data_bloqueio TEXT,
            data_execucao TEXT
        )
    """)
    cursor.execute("SELECT nome, data_bloqueio, data_execucao FROM bloqueios")
    historico_bloqueios = cursor.fetchall()
    conn.close()
    historico_bloqueios = [
        item for item in historico_bloqueios
        if item[0] in nomes_filtrados
    ]

    nomes_meses = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]
    ferias_por_mes = []
    bloqueios_por_mes = []
    for mes in range(1, 13):
        inicio_mes = datetime(ano_selecionado, mes, 1).date()
        fim_mes = (
            datetime(ano_selecionado + 1, 1, 1).date()
            if mes == 12
            else datetime(ano_selecionado, mes + 1, 1).date()
        ) - timedelta(days=1)
        ferias_por_mes.append(sum(
            1 for item in registros
            if item["inicio"] <= fim_mes and item["fim"] >= inicio_mes
        ))
        bloqueios_por_mes.append(sum(
            1 for _, data_bloqueio, _ in historico_bloqueios
            if str(data_bloqueio).startswith(f"{ano_selecionado}-{mes:02d}")
        ))

    departamentos_periodo = {}
    for item in registros_periodo:
        departamentos_periodo[item["departamento"]] = (
            departamentos_periodo.get(item["departamento"], 0) + 1
        )
    departamentos_ordenados = sorted(
        departamentos_periodo.items(),
        key=lambda item: (-item[1], normalizar_texto(item[0]))
    )

    status_contagem = {
        "active": sum(1 for item in colaboradores if item["status_classe"] == "active"),
        "scheduled": sum(1 for item in colaboradores if item["status_classe"] == "scheduled"),
        "completed": sum(1 for item in colaboradores if item["status_classe"] == "completed")
    }
    proximos_7_dias = [
        item for item in registros
        if hoje < item["inicio"] <= hoje + timedelta(days=7)
    ]
    proximos_bloqueios = [
        item for item in registros
        if item["bloqueio"] and item["bloqueio"] >= hoje
    ]
    retornos = [
        item for item in registros
        if item["fim"] >= hoje
    ]
    retornos_7_dias = [
        item for item in registros
        if hoje < item["fim"] <= hoje + timedelta(days=7)
    ]
    duracoes_regulares = [
        item["dias"] for item in registros_periodo
        if 1 <= item["dias"] <= 60
    ]
    maior_periodo = max(duracoes_regulares, default=0)
    departamentos_informados = [
        item for item in departamentos_ordenados
        if item[0] != "Não informado"
    ]
    mais_impactado = (
        departamentos_informados[0][0]
        if departamentos_informados else "Sem dados"
    )
    maior_volume = max(ferias_por_mes, default=0)
    mes_maior_volume = (
        nomes_meses[ferias_por_mes.index(maior_volume)]
        if maior_volume else "Sem dados"
    )
    percentual_em_ferias = (
        round((status_contagem["active"] / len(colaboradores)) * 100)
        if colaboradores else 0
    )
    caminho = planilha_mais_recente()
    ultima_atualizacao = (
        datetime.fromtimestamp(os.path.getmtime(caminho)).strftime("%d/%m/%Y")
        if caminho else "Sem importação"
    )

    periodos_disponiveis = set()
    for item in registros:
        cursor_mes = item["inicio"].replace(day=1)
        fim_cursor = item["fim"].replace(day=1)
        while cursor_mes <= fim_cursor:
            periodos_disponiveis.add(cursor_mes.strftime("%Y-%m"))
            cursor_mes = (
                cursor_mes.replace(year=cursor_mes.year + 1, month=1)
                if cursor_mes.month == 12
                else cursor_mes.replace(month=cursor_mes.month + 1)
            )
    periodos_disponiveis.add(periodo)

    return {
        "agora": agora,
        "periodo": periodo,
        "periodo_nome": f"{nomes_meses[mes_selecionado - 1]} de {ano_selecionado}",
        "departamento_filtro": departamento_filtro,
        "status_filtro": status_filtro,
        "pesquisa": pesquisa,
        "departamentos": departamentos,
        "periodos_disponiveis": [
            {
                "valor": valor,
                "rotulo": f"{nomes_meses[int(valor[5:7]) - 1]} {valor[:4]}"
            }
            for valor in sorted(periodos_disponiveis, reverse=True)
        ],
        "registros": registros_periodo,
        "total_colaboradores": len(colaboradores),
        "em_ferias": status_contagem["active"],
        "bloqueados": len(historico_bloqueios),
        "proximos": len(proximos_7_dias),
        "proximos_7_dias": len(proximos_7_dias),
        "retornos_7_dias": len(retornos_7_dias),
        "percentual_em_ferias": percentual_em_ferias,
        "mes_maior_volume": mes_maior_volume,
        "maior_volume": maior_volume,
        "meses": nomes_meses,
        "ferias_por_mes": ferias_por_mes,
        "bloqueios_por_mes": bloqueios_por_mes,
        "departamento_labels": [item[0] for item in departamentos_informados],
        "departamento_valores": [item[1] for item in departamentos_informados],
        "status_valores": [
            status_contagem["active"],
            status_contagem["scheduled"],
            status_contagem["completed"]
        ],
        "maior_periodo": maior_periodo,
        "proximo_bloqueio": min(
            (item["bloqueio"] for item in proximos_bloqueios),
            default=None
        ),
        "proximo_retorno": min(
            (item["fim"] for item in retornos),
            default=None
        ),
        "mais_impactado": mais_impactado,
        "ultima_atualizacao": ultima_atualizacao
    }


def obter_resumo_sistema():
    agora = datetime.now()
    hoje = agora.date()
    colaboradores = obter_colaboradores()
    caminhos = planilhas_importadas()
    caminho = planilha_mais_recente()
    bloqueios_hoje = 0
    inconsistencias = 0
    if caminhos:
        try:
            df = carregar_planilhas(caminhos)
            bloqueios_hoje = int(
                (df["Data de Bloqueio"].dt.date == hoje).sum()
            )
            inconsistencias = int(
                df["Inicio"].isna().sum()
                + df["Fim"].isna().sum()
                + ((df["Fim"] < df["Inicio"]) & df["Fim"].notna() & df["Inicio"].notna()).sum()
                + df.duplicated(subset=["Nome", "Inicio", "Fim"]).sum()
            )
        except (ValueError, OSError, KeyError):
            inconsistencias = 1

    periodos = [
        periodo
        for colaborador in colaboradores
        for periodo in colaborador["periodos"]
    ]
    retornos_hoje = sum(1 for item in periodos if item["fim"] == hoje)
    ferias_semana = sum(
        1 for item in periodos
        if hoje <= item["inicio"] <= hoje + timedelta(days=7)
    )
    proximos_bloqueios = sorted(
        item["bloqueio"] for item in periodos
        if item["bloqueio"] and item["bloqueio"] >= hoje
    )
    if agora.hour < 12:
        saudacao = "Bom dia"
    elif agora.hour < 18:
        saudacao = "Boa tarde"
    else:
        saudacao = "Boa noite"

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bloqueios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT,
            data_bloqueio TEXT,
            data_execucao TEXT
        )
    """)
    bloqueados = cursor.execute("SELECT COUNT(*) FROM bloqueios").fetchone()[0]
    conn.close()

    return {
        "saudacao": saudacao,
        "total_colaboradores": len(colaboradores),
        "em_ferias": sum(1 for item in colaboradores if item["status_classe"] == "active"),
        "bloqueados": bloqueados,
        "bloqueios_hoje": bloqueios_hoje,
        "retornos_hoje": retornos_hoje,
        "ferias_semana": ferias_semana,
        "inconsistencias": inconsistencias,
        "proximo_bloqueio": proximos_bloqueios[0] if proximos_bloqueios else None,
        "planilha_nome": os.path.basename(caminho) if caminho else "Nenhuma planilha importada",
        "ultima_importacao_hora": (
            datetime.fromtimestamp(os.path.getmtime(caminho)).strftime("%H:%M")
            if caminho else "--:--"
        ),
        "ultima_importacao_data": (
            datetime.fromtimestamp(os.path.getmtime(caminho)).strftime("%d/%m/%Y")
            if caminho else "Sem importação"
        )
    }


def obter_timeline_sistema(limite=6):
    inicializar_tabelas_sistema()
    conn = sqlite3.connect(DATABASE_PATH)
    registros = conn.execute(
        """
        SELECT data_hora, acao, detalhe
        FROM auditoria
        ORDER BY datetime(data_hora) DESC
        LIMIT ?
        """,
        (limite,)
    ).fetchall()
    conn.close()
    timeline = []
    for data_hora, acao, detalhe in registros:
        try:
            momento = datetime.fromisoformat(data_hora)
        except (TypeError, ValueError):
            momento = datetime.now()
        timeline.append({
            "hora": momento.strftime("%H:%M"),
            "data": momento.strftime("%d/%m"),
            "acao": acao,
            "detalhe": detalhe or "Ação registrada no sistema",
            "tipo": (
                "success" if "Importou" in acao
                else "purple" if "PDF" in acao
                else "warning" if "bloqueio" in acao.lower()
                else "blue"
            )
        })
    return timeline


def obter_inteligencia_sistema():
    agora = datetime.now()
    hoje = agora.date()
    amanha = hoje + timedelta(days=1)
    colaboradores = obter_colaboradores()
    periodos = [
        {**periodo, "nome": colaborador["nome"],
         "departamento": colaborador["departamento"]}
        for colaborador in colaboradores
        for periodo in colaborador["periodos"]
    ]
    quem_sai = sorted(
        [item for item in periodos if hoje <= item["inicio"] <= hoje + timedelta(days=7)],
        key=lambda item: item["inicio"]
    )
    quem_volta = sorted(
        [item for item in periodos if hoje <= item["fim"] <= hoje + timedelta(days=7)],
        key=lambda item: item["fim"]
    )

    inicializar_tabelas_sistema()
    conn = sqlite3.connect(DATABASE_PATH)
    bloqueios = conn.execute(
        "SELECT nome, data_bloqueio FROM bloqueios"
    ).fetchall() if conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='bloqueios'"
    ).fetchone() else []
    realizados = {(normalizar_texto(nome), data) for nome, data in bloqueios}
    pendentes = [
        item for item in periodos
        if item["bloqueio"] and item["bloqueio"] <= hoje
        and (
            normalizar_texto(item["nome"]),
            str(item["bloqueio"])
        ) not in realizados
    ]
    hoje_iso = hoje.isoformat()
    acoes_hoje = conn.execute(
        """
        SELECT acao, resultado FROM auditoria
        WHERE substr(data_hora, 1, 10) = ?
        """,
        (hoje_iso,)
    ).fetchall()
    conn.close()

    importacoes = sum(
        1 for acao, resultado in acoes_hoje
        if "Importou" in acao and resultado == "Sucesso"
    )
    erros = sum(1 for _, resultado in acoes_hoje if resultado == "Falha")
    bloqueios_hoje = sum(
        1 for item in periodos if item["bloqueio"] == hoje
    )
    ferias_amanha = sum(1 for item in periodos if item["inicio"] == amanha)
    retornos_hoje = sum(1 for item in periodos if item["fim"] == hoje)
    ferias_hoje = sum(
        1 for item in periodos if item["inicio"] <= hoje <= item["fim"]
    )

    departamentos = {}
    for colaborador in colaboradores:
        nome = colaborador["departamento"]
        departamentos[nome] = departamentos.get(nome, 0) + 1
    departamentos_lista = sorted(
        departamentos.items(), key=lambda item: item[1], reverse=True
    )

    caminhos = planilhas_importadas()
    caminho = planilha_mais_recente()
    vendas = []
    if caminhos:
        try:
            df = carregar_planilhas(caminhos)
            coluna_venda = encontrar_coluna(
                df, ["abono", "venda ferias", "vendeu ferias", "dias vendidos"]
            )
            if coluna_venda:
                for _, linha in df.iterrows():
                    valor = str(linha.get(coluna_venda, "")).strip()
                    if valor and valor.lower() not in {"nan", "não", "nao", "0", "0.0"}:
                        vendas.append({
                            "nome": str(linha["Nome"]),
                            "valor": valor
                        })
        except (ValueError, OSError):
            pass

    total = len(colaboradores)
    em_ferias = sum(
        1 for item in colaboradores if item["status_classe"] == "active"
    )
    ocupacao = round((em_ferias / total) * 100) if total else 0
    devidos = sum(1 for item in periodos if item["bloqueio"] and item["bloqueio"] <= hoje)
    execucao = round(((devidos - len(pendentes)) / devidos) * 100) if devidos else 100
    alertas = [
        {
            "nivel": "danger", "icone": "lock-keyhole",
            "titulo": f"{bloqueios_hoje} bloqueio(s) hoje",
            "descricao": "Acessos programados para bloqueio.",
            "href": "/dashboard#alertas"
        },
        {
            "nivel": "warning", "icone": "calendar-clock",
            "titulo": f"{ferias_amanha} férias amanhã",
            "descricao": "Colaboradores com saída programada.",
            "href": "/calendario"
        },
        {
            "nivel": "info", "icone": "log-in",
            "titulo": f"{retornos_hoje} retorno(s) hoje",
            "descricao": "Colaboradores com retorno programado.",
            "href": "/calendario"
        },
        {
            "nivel": "success", "icone": "file-check-2",
            "titulo": "Importação realizada" if caminho else "Sem importação",
            "descricao": (
                datetime.fromtimestamp(os.path.getmtime(caminho)).strftime(
                    "Base atualizada em %d/%m às %H:%M."
                ) if caminho else "Nenhuma base disponível."
            ),
            "href": "/importar"
        },
        {
            "nivel": "info", "icone": "database-backup",
            "titulo": "Backup executado" if listar_backups() else "Backup pendente",
            "descricao": (
                datetime.fromtimestamp(os.path.getmtime(listar_backups()[0])).strftime(
                    "Última cópia em %d/%m às %H:%M."
                ) if listar_backups() else "Crie a primeira cópia de segurança."
            ),
            "href": "/configuracoes#backup"
        }
    ]
    return {
        "alertas": alertas,
        "operacoes": {
            "bloqueios": bloqueios_hoje,
            "retornos": retornos_hoje,
            "ferias": ferias_hoje,
            "importacoes": importacoes,
            "erros": erros
        },
        "executivo": {
            "funcionarios": total, "em_ferias": em_ferias,
            "ocupacao": ocupacao, "bloqueios": bloqueios_hoje,
            "execucao": max(0, execucao),
            "departamentos": departamentos_lista[:6]
        },
        "rh": {
            "quem_sai": quem_sai[:12],
            "quem_volta": quem_volta[:12],
            "vendas": vendas[:12],
            "atrasados": pendentes[:12],
            "pendentes": len(pendentes)
        },
        "ti": {
            "importacoes": importacoes,
            "logs": len(acoes_hoje),
            "backups": len(listar_backups()),
            "erros": erros,
            "saude": obter_saude_sistema()
        }
    }


@app.context_processor
def contexto_notificacoes():
    hoje = datetime.now().date()
    amanha = hoje + timedelta(days=1)
    notificacoes = []
    caminhos = planilhas_importadas()
    configuracoes = carregar_configuracoes()

    if configuracoes.get("notificacoes_ativas") != "1":
        return {
            "notificacoes_topo": [],
            "notificacoes_total": 0,
            "versao_sistema": configuracoes.get("versao", "2.0"),
            "ultima_atualizacao_sistema": "27/07/2026",
            "mostrar_popup_notificacoes": False,
            "som_notificacoes_ativo": False,
            "notificacao_em_teste": False
        }

    if caminhos:
        caminho = caminhos[-1]
        try:
            df = carregar_planilhas(caminhos)
            linhas_bloqueio = df[df["Data de Bloqueio"].dt.date == hoje]
            linhas_amanha = df[df["Inicio"].dt.date == amanha]
            hora_rotina = configuracoes["hora_rotina"]
            if len(linhas_bloqueio):
                notificacoes.append({
                    "tipo": "danger",
                    "icone": "lock-keyhole",
                    "rotulo": "Bloquear hoje",
                    "titulo": f"{len(linhas_bloqueio)} colaborador(es) precisam ser bloqueados hoje.",
                    "descricao": "Usuários aguardando bloqueio de acesso.",
                    "href": "/dashboard#alertas",
                    "itens": [
                        {"nome": str(item["Nome"]), "meta": hora_rotina}
                        for _, item in linhas_bloqueio.head(5).iterrows()
                    ]
                })
            if len(linhas_amanha):
                notificacoes.append({
                    "tipo": "warning",
                    "icone": "calendar-clock",
                    "rotulo": "Próximas férias",
                    "titulo": f"{len(linhas_amanha)} colaborador(es) iniciam férias amanhã.",
                    "descricao": "Confira os colaboradores programados.",
                    "href": "/calendario",
                    "itens": [
                        {"nome": str(item["Nome"]), "meta": amanha.strftime("%d/%m")}
                        for _, item in linhas_amanha.head(5).iterrows()
                    ]
                })
        except (ValueError, OSError, KeyError):
            pass

        if configuracoes.get("notificar_importacao") == "1":
            notificacoes.append({
                "tipo": "success",
                "icone": "file-check-2",
                "rotulo": "Importação",
                "titulo": "Nova planilha importada.",
                "descricao": datetime.fromtimestamp(
                    os.path.getmtime(caminho)
                ).strftime("Planilha atualizada em %d/%m às %H:%M."),
                "href": "/dashboard",
                "itens": [{
                    "nome": os.path.basename(caminho),
                    "meta": datetime.fromtimestamp(
                        os.path.getmtime(caminho)
                    ).strftime("%H:%M")
                }]
            })

    backups = listar_backups()
    if backups:
        momento_backup = datetime.fromtimestamp(os.path.getmtime(backups[0]))
        notificacoes.append({
            "tipo": "success",
            "icone": "database-backup",
            "rotulo": "Backup concluído",
            "titulo": "Os dados do sistema estão protegidos.",
            "descricao": momento_backup.strftime(
                "Último backup em %d/%m/%Y às %H:%M."
            ),
            "href": "/configuracoes#backup",
            "itens": []
        })

    if request.args.get("teste_notificacao") == "1":
        notificacoes.insert(0, {
            "tipo": "success",
            "icone": "badge-check",
            "rotulo": "Notificação de teste",
            "titulo": "Sistema funcionando corretamente",
            "descricao": "A central de notificações está pronta para uso.",
            "href": "/configuracoes",
            "itens": []
        })

    return {
        "notificacoes_topo": notificacoes,
        "notificacoes_total": len(notificacoes),
        "versao_sistema": configuracoes.get("versao", "2.0"),
        "ultima_atualizacao_sistema": "27/07/2026",
        "mostrar_popup_notificacoes": configuracoes.get("mostrar_popup") == "1",
        "som_notificacoes_ativo": configuracoes.get("som_notificacoes") == "1",
        "notificacao_em_teste": request.args.get("teste_notificacao") == "1"
    }

# ======================
# 🏠 HOME
# ======================
@app.route("/")
def index():
    agora = datetime.now()
    inteligencia = obter_inteligencia_sistema()
    return render_template(
        "inicio.html",
        resumo=obter_resumo_sistema(),
        timeline=obter_timeline_sistema(8),
        inteligencia=inteligencia,
        saude_sistema=obter_saude_sistema(),
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y")
    )


@app.route("/alertas")
def centro_alertas():
    agora = datetime.now()
    return render_template(
        "inteligencia_painel.html",
        painel="alertas",
        dados=obter_inteligencia_sistema(),
        timeline=obter_timeline_sistema(8),
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y")
    )


@app.route("/operacoes")
def centro_operacoes():
    agora = datetime.now()
    return render_template(
        "inteligencia_painel.html",
        painel="operacoes",
        dados=obter_inteligencia_sistema(),
        timeline=obter_timeline_sistema(10),
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y")
    )


@app.route("/dashboard/<tipo>")
def dashboard_especializado(tipo):
    if tipo not in {"executivo", "rh", "ti"}:
        return redirect(url_for("dashboard"))
    agora = datetime.now()
    return render_template(
        "inteligencia_painel.html",
        painel=tipo,
        dados=obter_inteligencia_sistema(),
        timeline=obter_timeline_sistema(10),
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y")
    )


@app.route("/importar")
def importar():
    agora = datetime.now()
    return render_template(
        "index.html",
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y")
    )


@app.route("/api/importacao/validar", methods=["POST"])
def validar_importacao():
    inicio_validacao = time.perf_counter()
    arquivo = request.files.get("arquivo")
    modo = request.form.get("modo", "definitivo").strip().lower()
    simulacao = modo == "simulacao"
    if not arquivo or not arquivo.filename:
        return {
            "ok": False,
            "mensagem": "Selecione uma planilha para iniciar a análise."
        }, 400

    nome_original = os.path.basename(arquivo.filename)
    _, extensao = os.path.splitext(nome_original)
    extensao = extensao.lower()
    if extensao not in {".xlsx", ".xls"}:
        return {
            "ok": False,
            "mensagem": "Formato inválido. Envie um arquivo Excel .xlsx ou .xls."
        }, 400

    limpar_validacoes_temporarias()
    token = uuid.uuid4().hex
    os.makedirs(VALIDACOES_DIR, exist_ok=True)
    caminho = os.path.join(VALIDACOES_DIR, f"{token}{extensao}")
    arquivo.save(caminho)
    motor = obter_motor_importacao(readonly=simulacao)
    ip = obter_ip_requisicao()

    try:
        resultado = motor.analyze(
            caminho,
            planilha_mais_recente(),
            actor_user="T.Costa",
            actor_ip=ip,
            dry_run=simulacao
        )
        resultado["arquivo"] = nome_original
        duracao = round(time.perf_counter() - inicio_validacao, 3)
        metadados = {
            "token": token,
            "arquivo": nome_original,
            "extensao": extensao,
            "modo": "simulacao" if simulacao else "definitivo",
            "criado_em": datetime.now().isoformat(timespec="seconds"),
            "duracao_validacao": duracao,
            "resultado": resultado
        }
        with open(
            caminho_metadados_validacao(token), "w", encoding="utf-8"
        ) as arquivo_metadados:
            json.dump(
                metadados,
                arquivo_metadados,
                ensure_ascii=False,
                indent=2
            )
    except PluginExecutionError as erro:
        if os.path.exists(caminho):
            os.remove(caminho)
        if not simulacao:
            motor.record_failure(
                arquivo=nome_original,
                quantidade=0,
                tempo_segundos=round(
                    time.perf_counter() - inicio_validacao,
                    3
                ),
                erros=1,
                usuario="T.Costa",
                ip=ip,
                mensagem=str(erro)
            )
        return {
            "ok": False,
            "mensagem": (
                "A validação foi interrompida por uma etapa adicional "
                "configurada. Consulte os logs do sistema."
            )
        }, 400
    except Exception as erro:
        # Arquivos externos podem falhar em diferentes leitores do Excel.
        if os.path.exists(caminho):
            os.remove(caminho)
        if not simulacao:
            motor.record_failure(
                arquivo=nome_original,
                quantidade=0,
                tempo_segundos=round(
                    time.perf_counter() - inicio_validacao,
                    3
                ),
                erros=1,
                usuario="T.Costa",
                ip=ip,
                mensagem=f"Falha ao ler o arquivo: {erro}"
            )
        return {
            "ok": False,
            "mensagem": (
                "Não foi possível analisar a planilha. "
                "Verifique se o arquivo está íntegro."
            )
        }, 400

    if not simulacao:
        motor.record_validation(
            resultado=resultado,
            tempo_segundos=duracao,
            usuario="T.Costa",
            ip=ip
        )
    return {
        "ok": True,
        "token": token,
        "duracao_segundos": duracao,
        "validacao": resultado
    }


@app.route("/api/importacao/mapeamento", methods=["POST"])
def confirmar_mapeamento_importacao():
    inicio_validacao = time.perf_counter()
    payload = request.get_json(silent=True) or {}
    token = str(payload.get("token", "")).strip().lower()
    mapeamento = payload.get("mapeamento")
    metadados = carregar_validacao_temporaria(token)
    if not metadados:
        return {
            "ok": False,
            "mensagem": "A validação expirou. Selecione a planilha novamente."
        }, 404
    if not isinstance(mapeamento, dict) or not mapeamento:
        return {
            "ok": False,
            "mensagem": "Informe o mapeamento de colunas para continuar."
        }, 400

    simulacao = metadados.get("modo") == "simulacao"
    motor = obter_motor_importacao(readonly=simulacao)
    resultado = motor.analyze(
        metadados["caminho"],
        planilha_mais_recente(),
        confirmed_mapping=mapeamento,
        profile_name=payload.get("nome_perfil"),
        profile_origin=payload.get("origem"),
        actor_user="T.Costa",
        actor_ip=obter_ip_requisicao(),
        dry_run=simulacao
    )
    resultado["arquivo"] = metadados["arquivo"]
    duracao = round(time.perf_counter() - inicio_validacao, 3)
    metadados["resultado"] = resultado
    metadados["duracao_validacao"] = duracao
    metadados["mapeamento_confirmado_em"] = datetime.now().isoformat(
        timespec="seconds"
    )
    with open(
        metadados["caminho_metadados"], "w", encoding="utf-8"
    ) as arquivo_metadados:
        json.dump(
            {
                chave: valor
                for chave, valor in metadados.items()
                if chave not in {"caminho", "caminho_metadados"}
            },
            arquivo_metadados,
            ensure_ascii=False,
            indent=2
        )
    if not simulacao:
        motor.record_validation(
            resultado=resultado,
            tempo_segundos=duracao,
            usuario="T.Costa",
            ip=obter_ip_requisicao()
        )
    return {
        "ok": True,
        "token": token,
        "duracao_segundos": duracao,
        "validacao": resultado
    }


@app.route("/api/importacao/perfis", methods=["GET"])
def listar_perfis_importacao():
    motor = obter_motor_importacao()
    incluir_inativos = request.args.get("incluir_inativos") == "1"
    return {
        "ok": True,
        "perfis": motor.profile_store.list(
            active_only=not incluir_inativos
        )
    }


@app.route("/api/importacao/atualizacoes", methods=["GET"])
def listar_atualizacoes_importacao():
    motor = obter_motor_importacao()
    return {
        "ok": True,
        "modulos": motor.dashboard_updater.status()
    }


@app.route("/api/importacao/plugins", methods=["GET"])
def listar_plugins_importacao():
    motor = obter_motor_importacao()
    return {
        "ok": True,
        "plugins": motor.plugin_manager.list(),
        "erros_carregamento": motor.plugin_manager.load_errors
    }


@app.route("/api/importacao/plugins/<string:nome>", methods=["PATCH"])
def configurar_plugin_importacao(nome):
    payload = request.get_json(silent=True) or {}
    ativo = payload.get("ativo") if "ativo" in payload else None
    prioridade = (
        payload.get("prioridade")
        if "prioridade" in payload else None
    )
    if ativo is not None and not isinstance(ativo, bool):
        return {
            "ok": False,
            "mensagem": "O campo ativo deve ser verdadeiro ou falso."
        }, 400
    motor = obter_motor_importacao()
    try:
        alterado = motor.plugin_manager.configure(
            nome,
            enabled=ativo,
            priority=prioridade
        )
    except ValueError as erro:
        return {"ok": False, "mensagem": str(erro)}, 400
    if not alterado:
        return {
            "ok": False,
            "mensagem": "Plugin de importação não encontrado."
        }, 404
    registrar_auditoria(
        "Configurou plugin de importação",
        (
            f"Plugin: {nome} · "
            f"Ativo: {ativo if ativo is not None else 'inalterado'} · "
            f"Prioridade: "
            f"{prioridade if prioridade is not None else 'inalterada'}"
        ),
        usuario="T.Costa",
        ip=obter_ip_requisicao()
    )
    return {"ok": True}


@app.route("/api/importacao/perfis/<int:perfil_id>", methods=["PATCH"])
def atualizar_perfil_importacao(perfil_id):
    payload = request.get_json(silent=True) or {}
    motor = obter_motor_importacao()
    alterado = False
    detalhes = []
    try:
        if "nome" in payload:
            alterado = motor.profile_store.rename(
                perfil_id,
                payload["nome"]
            ) or alterado
            detalhes.append(f"Nome: {str(payload['nome']).strip()}")
        if "ativo" in payload:
            if not isinstance(payload["ativo"], bool):
                raise ValueError(
                    "O campo ativo deve ser verdadeiro ou falso."
                )
            ativo = payload["ativo"]
            alterado = motor.profile_store.set_active(
                perfil_id,
                ativo
            ) or alterado
            detalhes.append("Ativo" if ativo else "Inativo")
    except ValueError as erro:
        return {"ok": False, "mensagem": str(erro)}, 400
    if not alterado:
        return {
            "ok": False,
            "mensagem": "Perfil de importação não encontrado."
        }, 404
    registrar_auditoria(
        "Atualizou perfil de importação",
        f"Perfil {perfil_id} · {' · '.join(detalhes)}",
        usuario="T.Costa",
        ip=obter_ip_requisicao()
    )
    return {"ok": True}

# ======================
# 📤 UPLOAD
# ======================
@app.route("/upload", methods=["POST"])
def upload():
    inicio_importacao = time.perf_counter()
    configuracoes_importacao = carregar_configuracoes()
    motor = obter_motor_importacao()
    ip = obter_ip_requisicao()
    token = request.form.get("validacao_token", "").strip().lower()
    metadados = carregar_validacao_temporaria(token)
    arquivo = request.files.get("arquivo")

    if metadados and metadados.get("modo") == "simulacao":
        return render_template(
            "resultado.html",
            usuarios=[],
            erro=(
                "Uma simulação não pode ser confirmada como importação. "
                "Selecione o modo definitivo e analise a planilha novamente."
            )
        ), 400

    if metadados:
        caminho = metadados["caminho"]
        nome_arquivo = metadados["arquivo"]
        caminho_metadados = metadados["caminho_metadados"]
    elif arquivo and arquivo.filename:
        nome_arquivo = os.path.basename(arquivo.filename)
        _, extensao = os.path.splitext(nome_arquivo)
        extensao = extensao.lower()
        if extensao not in {".xlsx", ".xls"}:
            return render_template(
                "resultado.html",
                usuarios=[],
                erro="Formato inválido. Envie uma planilha Excel .xlsx ou .xls."
            )
        token = uuid.uuid4().hex
        os.makedirs(VALIDACOES_DIR, exist_ok=True)
        caminho = os.path.join(VALIDACOES_DIR, f"{token}{extensao}")
        caminho_metadados = None
        arquivo.save(caminho)
    else:
        return render_template(
            "resultado.html",
            usuarios=[],
            erro="Selecione e analise uma planilha antes de importar."
        )

    caminho_anterior = planilha_mais_recente()
    try:
        resultado_validacao = motor.analyze(caminho, caminho_anterior)
        resultado_validacao["arquivo"] = nome_arquivo
        if not resultado_validacao["pronta"]:
            primeiros_erros = "; ".join(
                f"Linha {item['linha']}: {item['mensagem']}"
                for item in resultado_validacao["erros"][:3]
            )
            raise ValueError(
                f"Foram encontrados {resultado_validacao['total_erros']} erro(s). "
                f"{primeiros_erros}"
            )
        df = carregar_planilha(
            caminho,
            resultado_validacao["mapeamento"]["colunas"]
        )
    except Exception as erro:
        # A confirmação nunca deve expor um erro técnico ao usuário.
        if os.path.exists(caminho):
            os.remove(caminho)
        if caminho_metadados and os.path.exists(caminho_metadados):
            os.remove(caminho_metadados)
        motor.record_failure(
            arquivo=nome_arquivo,
            quantidade=resultado_validacao.get("total_registros", 0)
            if "resultado_validacao" in locals() else 0,
            tempo_segundos=round(
                time.perf_counter() - inicio_importacao,
                3
            ),
            erros=resultado_validacao.get("total_erros", 1)
            if "resultado_validacao" in locals() else 1,
            usuario="T.Costa",
            ip=ip,
            mensagem=f"Falha na validação: {erro}"
        )
        return render_template(
            "resultado.html",
            usuarios=[],
            erro=str(erro)
        )

    try:
        caminho_backup = motor.prepare_import(
            operation_id=resultado_validacao.get("operacao_id"),
            filename=nome_arquivo,
            user="T.Costa",
            ip=ip
        )
    except PluginExecutionError as erro:
        motor.record_failure(
            arquivo=nome_arquivo,
            quantidade=resultado_validacao["total_registros"],
            tempo_segundos=round(
                time.perf_counter() - inicio_importacao,
                3
            ),
            erros=1,
            usuario="T.Costa",
            ip=ip,
            mensagem=str(erro),
            operation_id=resultado_validacao.get("operacao_id")
        )
        return render_template(
            "resultado.html",
            usuarios=[],
            erro=(
                "A importação foi interrompida por uma etapa adicional "
                "configurada. Consulte os logs do sistema."
            )
        )
    except (OSError, sqlite3.Error) as erro:
        motor.record_failure(
            arquivo=nome_arquivo,
            quantidade=resultado_validacao["total_registros"],
            tempo_segundos=round(
                time.perf_counter() - inicio_importacao,
                3
            ),
            erros=1,
            usuario="T.Costa",
            ip=ip,
            mensagem=f"Backup obrigatório não realizado: {erro}"
        )
        return render_template(
            "resultado.html",
            usuarios=[],
            erro=(
                "A importação foi interrompida porque o backup obrigatório "
                "não pôde ser criado."
            )
        )

    pasta_planilhas = obter_pasta_planilhas()
    caminho_final = os.path.join(pasta_planilhas, nome_arquivo)

    if configuracoes_importacao.get("substituir_planilha") == "1":
        for nome_antigo in os.listdir(pasta_planilhas):
            caminho_antigo = os.path.join(pasta_planilhas, nome_antigo)
            if (
                caminho_antigo != caminho_final
                and os.path.isfile(caminho_antigo)
                and nome_antigo.lower().endswith((".xlsx", ".xls"))
            ):
                os.remove(caminho_antigo)
    os.replace(caminho, caminho_final)
    if caminho_metadados and os.path.exists(caminho_metadados):
        os.remove(caminho_metadados)

    dados = []

    for _, row in df.iterrows():

        try:

            nome = str(row["Nome"]).strip()

            bloqueio = row["Data de Bloqueio"]

            if nome == "nan":
                continue

            dados.append({
                "Nome": nome,
                "Bloqueio": bloqueio.date()
                if pd.notnull(bloqueio)
                else None
            })

        except (KeyError, TypeError, ValueError):
            continue

    hoje = datetime.now().date()

    usuarios = []
    bloqueios_registrados = []

    conn = sqlite3.connect(DATABASE_PATH)

    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bloqueios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT,
            data_bloqueio TEXT,
            data_execucao TEXT,
            UNIQUE(nome, data_bloqueio)
        )
    """)

    for d in dados:

        if d["Bloqueio"] is None:
            continue

        if d["Bloqueio"] == hoje:
            usuarios.append(d["Nome"])

        if d["Bloqueio"] <= hoje:
            cursor.execute("""
                INSERT OR IGNORE INTO bloqueios
                (nome, data_bloqueio, data_execucao)
                VALUES (?, ?, ?)
            """, (
                d["Nome"],
                str(d["Bloqueio"]),
                str(datetime.now())
            ))
            if cursor.rowcount:
                bloqueios_registrados.append(d["Nome"])

    conn.commit()

    conn.close()

    duracao_importacao = round(time.perf_counter() - inicio_importacao, 3)
    comparacao = resultado_validacao["comparacao"]
    motor.complete(
        caminho_arquivo=caminho_final,
        arquivo=nome_arquivo,
        registros=resultado_validacao["total_registros"],
        duracao_segundos=duracao_importacao,
        usuario="T.Costa",
        ip=ip,
        comparacao=comparacao,
        backup=caminho_backup,
        extra_payload={
            "bloqueios_registrados": bloqueios_registrados
        },
        operation_id=resultado_validacao.get("operacao_id")
    )

    if configuracoes_importacao.get("mostrar_resumo_importacao") == "1":
        return render_template(
            "resultado.html",
            usuarios=usuarios,
            importado=True
        )
    return redirect(url_for("dashboard", importado="1"))

# ======================
# 📊 DASHBOARD
# ======================
@app.route("/dashboard")
def dashboard():

    agora = datetime.now()
    data_atual = agora.strftime("%d/%m/%Y %H:%M")
    data_hoje = agora.strftime("%d/%m/%Y")
    resumo_inteligente = obter_resumo_sistema()
    timeline_sistema = obter_timeline_sistema()

    caminhos = planilhas_importadas()
    caminho = planilha_mais_recente()

    if not caminhos:
        return render_template(
            "dashboard.html",
            hoje=[],
            proximos=[],
            bloqueados=[],
            todos=[],
            em_ferias=[],
            bloqueios_hoje_total=0,
            labels=[],
            valores=[],
            data_atual=data_atual,
            data_hoje=data_hoje,
            resumo_inteligente=resumo_inteligente,
            timeline_sistema=timeline_sistema,
            planilha_nome="Nenhuma planilha importada",
            ultima_importacao="Sem importação"
        )

    try:
        df = carregar_planilhas(caminhos)
    except ValueError as erro:
        return str(erro)

    hoje = datetime.now().date()

    # 🔴 HOJE (inclui bloqueios marcados e quem inicia férias hoje)
    hoje_lista = df[
        (
            df["Data de Bloqueio"].dt.date == hoje
        ) |
        (
            df["Inicio"].dt.date == hoje
        )
    ]
    bloqueios_hoje_total = int(
        (df["Data de Bloqueio"].dt.date == hoje).sum()
    )

    # 🟡 PRÓXIMOS DIAS
    proximos = df[
        (
            df["Data de Bloqueio"].dt.date > hoje
        ) &
        (
            df["Data de Bloqueio"].dt.date
            <= hoje + pd.Timedelta(days=3)
        )
    ]

    em_ferias = df[
        (
            df["Inicio"].dt.date <= hoje
        ) &
        (
            df["Fim"].dt.date >= hoje
        )
    ]

    grafico = (
        df.dropna(subset=["Data de Bloqueio"])
        .groupby(df["Data de Bloqueio"].dt.strftime("%d/%m/%Y"))
        .size()
        .sort_index()
    )

    labels = grafico.index.tolist()
    valores = grafico.values.tolist()

    hoje_lista = formatar_datas(hoje_lista)
    proximos = formatar_datas(proximos)
    em_ferias = formatar_datas(em_ferias)
    df = formatar_datas(df)

    # 🟢 HISTÓRICO
    conn = sqlite3.connect(DATABASE_PATH)

    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bloqueios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT,
            data_bloqueio TEXT,
            data_execucao TEXT
        )
    """)

    cursor.execute(
        "SELECT nome FROM bloqueios"
    )

    bloqueados = [
        x[0]
        for x in cursor.fetchall()
    ]

    conn.close()

    planilha_nome = os.path.basename(caminho)
    ultima_importacao = datetime.fromtimestamp(
        os.path.getmtime(caminho)
    ).strftime("%d/%m/%Y %H:%M")

    return render_template(
        "dashboard.html",
        hoje=hoje_lista.to_dict(orient="records"),
        proximos=proximos.to_dict(orient="records"),
        bloqueados=bloqueados,
        todos=df.to_dict(orient="records"),
        em_ferias=em_ferias.to_dict(orient="records"),
        bloqueios_hoje_total=bloqueios_hoje_total,
        labels=labels,
        valores=valores,
        data_atual=data_atual,
        data_hoje=data_hoje,
        resumo_inteligente=resumo_inteligente,
        timeline_sistema=timeline_sistema,
        planilha_nome=planilha_nome,
        ultima_importacao=ultima_importacao
    )


@app.route("/detalhe/<secao>")
def detalhe(secao):

    caminhos = planilhas_importadas()

    if not caminhos:
        # para bloqueados ainda vamos buscar do banco
        if secao == "bloqueados":
            conn = sqlite3.connect(DATABASE_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT nome FROM bloqueios")
            bloqueados = [x[0] for x in cursor.fetchall()]
            conn.close()
            return render_template("detalhe.html", secao=secao, items=bloqueados)

        return render_template("detalhe.html", secao=secao, items=[])

    try:
        df = carregar_planilhas(caminhos)
    except ValueError:
        return render_template("detalhe.html", secao=secao, items=[])

    hoje = datetime.now().date()

    hoje_lista = df[
        (
            df["Data de Bloqueio"].dt.date == hoje
        ) |
        (
            df["Inicio"].dt.date == hoje
        )
    ]

    proximos = df[
        (
            df["Data de Bloqueio"].dt.date > hoje
        ) &
        (
            df["Data de Bloqueio"].dt.date
            <= hoje + pd.Timedelta(days=3)
        )
    ]

    df = formatar_datas(df)
    hoje_lista = formatar_datas(hoje_lista)
    proximos = formatar_datas(proximos)

    # histórico bloqueados
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS bloqueios (id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT, data_bloqueio TEXT, data_execucao TEXT)")
    cursor.execute("SELECT nome FROM bloqueios")
    bloqueados = [x[0] for x in cursor.fetchall()]
    conn.close()

    mapping = {
        'hoje': hoje_lista.to_dict(orient='records'),
        'proximos': proximos.to_dict(orient='records'),
        'bloqueados': bloqueados,
        'todos': df.to_dict(orient='records')
    }

    items = mapping.get(secao, [])

    return render_template('detalhe.html', secao=secao, items=items)


@app.route("/calendario")
def calendario():
    agora = datetime.now()
    mes_parametro = request.args.get("mes", agora.strftime("%Y-%m"))

    try:
        ano, mes = [int(parte) for parte in mes_parametro.split("-", 1)]
        primeiro_dia = datetime(ano, mes, 1).date()
    except (TypeError, ValueError):
        primeiro_dia = agora.date().replace(day=1)
        ano, mes = primeiro_dia.year, primeiro_dia.month

    if mes == 12:
        proximo_primeiro = datetime(ano + 1, 1, 1).date()
    else:
        proximo_primeiro = datetime(ano, mes + 1, 1).date()
    ultimo_dia = proximo_primeiro - timedelta(days=1)
    mes_anterior = (primeiro_dia - timedelta(days=1)).replace(day=1)

    ferias = []
    caminhos = planilhas_importadas()
    if caminhos:
        try:
            df = carregar_planilhas(caminhos)
            coluna_departamento = encontrar_coluna(
                df,
                ["departamento", "depto / setor", "depto", "setor"]
            )
            df_mes = df[
                (df["Inicio"].dt.date <= ultimo_dia)
                & (df["Fim"].dt.date >= primeiro_dia)
            ].copy()

            for indice, (_, linha) in enumerate(df_mes.sort_values("Inicio").iterrows()):
                inicio = linha["Inicio"].date()
                fim = linha["Fim"].date()
                if inicio <= agora.date() <= fim:
                    status = "Em férias"
                    status_classe = "active"
                elif inicio > agora.date():
                    status = "Programada"
                    status_classe = "scheduled"
                else:
                    status = "Concluída"
                    status_classe = "completed"

                nome = str(linha["Nome"])
                departamento = (
                    str(linha[coluna_departamento]).strip()
                    if coluna_departamento
                    and pd.notna(linha[coluna_departamento])
                    else "Não informado"
                )
                bloqueio = (
                    linha["Data de Bloqueio"].date()
                    if pd.notna(linha["Data de Bloqueio"])
                    else None
                )
                ferias.append({
                    "id": indice,
                    "nome": nome,
                    "departamento": departamento,
                    "inicio": inicio,
                    "fim": fim,
                    "bloqueio": bloqueio,
                    "inicio_formatado": inicio.strftime("%d/%m/%Y"),
                    "fim_formatado": fim.strftime("%d/%m/%Y"),
                    "bloqueio_formatado": bloqueio.strftime("%d/%m/%Y") if bloqueio else "-",
                    "periodo_curto": f"{inicio:%d/%m} - {fim:%d/%m}",
                    "status": status,
                    "status_classe": status_classe,
                    "cor": sum(ord(letra) for letra in nome) % 4
                })
        except ValueError:
            ferias = []

    calendario_base = calendar_module.Calendar(firstweekday=0)
    semanas = []
    for semana in calendario_base.monthdatescalendar(ano, mes):
        dias = []
        for data in semana:
            eventos = [
                evento for evento in ferias
                if evento["inicio"] <= data <= evento["fim"]
            ]
            dias.append({
                "data": data,
                "numero": data.day,
                "mes_atual": data.month == mes,
                "hoje": data == agora.date(),
                "eventos": eventos
            })
        semanas.append(dias)

    nomes_meses = [
        "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]
    agenda = sorted(
        ferias,
        key=lambda item: (
            item["inicio"] < agora.date(),
            item["inicio"]
        )
    )[:6]

    return render_template(
        "calendario.html",
        semanas=semanas,
        ferias=ferias,
        ferias_json=[
            {
                "id": item["id"],
                "nome": item["nome"],
                "departamento": item["departamento"],
                "inicio": item["inicio_formatado"],
                "fim": item["fim_formatado"],
                "bloqueio": item["bloqueio_formatado"],
                "status": item["status"],
                "status_classe": item["status_classe"],
                "cor": item["cor"]
            }
            for item in ferias
        ],
        agenda=agenda,
        mes_nome=nomes_meses[mes],
        ano=ano,
        mes_parametro=f"{ano}-{mes:02d}",
        mes_anterior=mes_anterior.strftime("%Y-%m"),
        proximo_mes=proximo_primeiro.strftime("%Y-%m"),
        ferias_hoje=sum(1 for item in ferias if item["status_classe"] == "active"),
        proximas=sum(1 for item in ferias if item["status_classe"] == "scheduled"),
        colaboradores=len({item["nome"] for item in ferias}),
        busca=request.args.get("busca", ""),
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y")
    )


def gerar_pdf_calendario(ferias, mes_nome, ano):
    preferencias = carregar_configuracoes()
    def pdf_texto(valor):
        texto = str(valor).encode("latin-1", "replace").decode("latin-1")
        return texto.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    linhas_por_pagina = 23
    paginas = [ferias[i:i + linhas_por_pagina] for i in range(0, len(ferias), linhas_por_pagina)] or [[]]
    objetos = {
        1: b"<< /Type /Catalog /Pages 2 0 R >>",
        3: b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>",
        4: b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold /Encoding /WinAnsiEncoding >>"
    }
    referencias_paginas = []
    proximo_objeto = 5

    for indice, pagina in enumerate(paginas, start=1):
        pagina_id, conteudo_id = proximo_objeto, proximo_objeto + 1
        proximo_objeto += 2
        referencias_paginas.append(f"{pagina_id} 0 R")
        comandos = [
            "0.032 0.184 0.357 rg 0 772 595 70 re f",
            "BT /F2 19 Tf 1 1 1 rg 42 806 Td (Calendario de Ferias) Tj ET",
            f"BT /F1 10 Tf 0.82 0.9 1 rg 42 787 Td ({pdf_texto(mes_nome)} de {ano}) Tj ET",
            "0.93 0.96 0.99 rg 35 734 525 28 re f",
            "BT /F2 8 Tf 0.18 0.25 0.34 rg 42 744 Td (COLABORADOR) Tj ET",
            "BT /F2 8 Tf 0.18 0.25 0.34 rg 205 744 Td (SETOR) Tj ET",
            "BT /F2 8 Tf 0.18 0.25 0.34 rg 325 744 Td (INICIO) Tj ET",
            "BT /F2 8 Tf 0.18 0.25 0.34 rg 385 744 Td (FIM) Tj ET",
            "BT /F2 8 Tf 0.18 0.25 0.34 rg 445 744 Td (BLOQUEIO) Tj ET",
            "BT /F2 8 Tf 0.18 0.25 0.34 rg 510 744 Td (STATUS) Tj ET"
        ]
        if preferencias.get("pdf_logo") == "1":
            comandos.append("BT /F2 9 Tf 0.45 0.72 1 rg 500 806 Td (FOKUS) Tj ET")
        if preferencias.get("pdf_data") == "1":
            comandos.append(
                f"BT /F1 7 Tf 0.55 0.62 0.7 rg 430 787 Td (Gerado em {datetime.now():%d/%m/%Y}) Tj ET"
            )
        y = 715
        for linha, item in enumerate(pagina):
            if linha % 2:
                comandos.append(f"0.97 0.98 0.99 rg 35 {y - 8} 525 27 re f")
            comandos.extend([
                f"BT /F1 8 Tf 0.16 0.22 0.3 rg 42 {y} Td ({pdf_texto(item['nome'])[:34]}) Tj ET",
                f"BT /F1 8 Tf 0.3 0.37 0.46 rg 205 {y} Td ({pdf_texto(item['departamento'])[:22]}) Tj ET",
                f"BT /F1 8 Tf 0.3 0.37 0.46 rg 325 {y} Td ({item['inicio_formatado'][:5]}) Tj ET",
                f"BT /F1 8 Tf 0.3 0.37 0.46 rg 385 {y} Td ({item['fim_formatado'][:5]}) Tj ET",
                f"BT /F1 8 Tf 0.3 0.37 0.46 rg 445 {y} Td ({item['bloqueio_formatado'][:5]}) Tj ET",
                f"BT /F2 7 Tf 0.07 0.45 0.38 rg 510 {y} Td ({pdf_texto(item['status'])[:10]}) Tj ET"
            ])
            y -= 28
        if preferencias.get("pdf_rodape") == "1":
            comandos.append("BT /F1 8 Tf 0.45 0.5 0.58 rg 35 28 Td (Grupo Fokus Logistica - Sistema de Controle de Ferias) Tj ET")
        if preferencias.get("pdf_numero_pagina") == "1":
            comandos.append(f"BT /F1 8 Tf 0.45 0.5 0.58 rg 520 28 Td ({indice}/{len(paginas)}) Tj ET")
        stream = "\n".join(comandos).encode("latin-1")
        objetos[conteudo_id] = f"<< /Length {len(stream)} >>\nstream\n".encode("ascii") + stream + b"\nendstream"
        objetos[pagina_id] = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
            f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents {conteudo_id} 0 R >>"
        ).encode("ascii")

    objetos[2] = f"<< /Type /Pages /Kids [{' '.join(referencias_paginas)}] /Count {len(referencias_paginas)} >>".encode("ascii")
    pdf = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0] * (max(objetos) + 1)
    for numero in range(1, max(objetos) + 1):
        offsets[numero] = len(pdf)
        pdf.extend(f"{numero} 0 obj\n".encode("ascii"))
        pdf.extend(objetos[numero])
        pdf.extend(b"\nendobj\n")
    xref = len(pdf)
    pdf.extend(f"xref\n0 {len(offsets)}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(f"trailer\n<< /Size {len(offsets)} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode("ascii"))
    return bytes(pdf)


@app.route("/calendario/exportar/pdf")
def exportar_calendario_pdf():
    agora = datetime.now()
    mes_parametro = request.args.get("mes", agora.strftime("%Y-%m"))
    try:
        ano, mes = [int(parte) for parte in mes_parametro.split("-", 1)]
        primeiro_dia = datetime(ano, mes, 1).date()
    except (TypeError, ValueError):
        primeiro_dia = agora.date().replace(day=1)
        ano, mes = primeiro_dia.year, primeiro_dia.month
    proximo_primeiro = (
        datetime(ano + 1, 1, 1).date()
        if mes == 12
        else datetime(ano, mes + 1, 1).date()
    )
    ultimo_dia = proximo_primeiro - timedelta(days=1)
    ferias = []
    for colaborador in obter_colaboradores():
        for periodo in colaborador["periodos"]:
            if periodo["inicio"] <= ultimo_dia and periodo["fim"] >= primeiro_dia:
                ferias.append({
                    "nome": colaborador["nome"],
                    "departamento": colaborador["departamento"],
                    "inicio_formatado": periodo["inicio_formatado"],
                    "fim_formatado": periodo["fim_formatado"],
                    "bloqueio_formatado": periodo["bloqueio_formatado"],
                    "status": (
                        "Em ferias"
                        if periodo["inicio"] <= agora.date() <= periodo["fim"]
                        else "Programada"
                        if periodo["inicio"] > agora.date()
                        else "Concluida"
                    )
                })
    nomes_meses = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    registrar_auditoria(
        "Gerou PDF",
        f"Calendário de férias · {nomes_meses[mes]}/{ano}"
    )
    arquivo = BytesIO(gerar_pdf_calendario(ferias, nomes_meses[mes], ano))
    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"calendario_ferias_{ano}_{mes:02d}.pdf",
        mimetype="application/pdf"
    )


def obter_colaboradores():
    caminhos = planilhas_importadas()
    if not caminhos:
        return []

    try:
        df = carregar_planilhas(caminhos)
    except ValueError:
        return []

    coluna_departamento = encontrar_coluna(df, ["departamento", "depto / setor", "depto", "setor"])
    coluna_cargo = encontrar_coluna(df, ["cargo", "funcao"])
    coluna_matricula = encontrar_coluna(df, ["matricula"])
    coluna_filial = encontrar_coluna(df, ["filial", "unidade"])

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS colaborador_perfis (
            nome TEXT PRIMARY KEY,
            departamento TEXT,
            cargo TEXT,
            matricula TEXT,
            filial TEXT,
            atualizado_em TEXT
        )
    """)
    cursor.execute("SELECT nome, departamento, cargo, matricula, filial FROM colaborador_perfis")
    substituicoes = {
        normalizar_texto(nome): {
            "departamento": departamento,
            "cargo": cargo,
            "matricula": matricula,
            "filial": filial
        }
        for nome, departamento, cargo, matricula, filial in cursor.fetchall()
    }
    conn.close()

    def texto_valido(valor, padrao="Não informado"):
        if valor is None or pd.isna(valor) or str(valor).strip() in ("", "nan"):
            return padrao
        texto = str(valor).strip()
        if texto.endswith(".0") and texto[:-2].isdigit():
            return texto[:-2]
        return texto

    grupos = {}
    hoje = datetime.now().date()
    for _, linha in df.iterrows():
        nome = str(linha["Nome"]).strip()
        chave = normalizar_texto(nome)
        if chave not in grupos:
            grupos[chave] = {
                "nome": nome,
                "departamento": texto_valido(linha.get(coluna_departamento) if coluna_departamento else None),
                "cargo": texto_valido(linha.get(coluna_cargo) if coluna_cargo else None),
                "matricula": texto_valido(linha.get(coluna_matricula) if coluna_matricula else None, "-"),
                "filial": texto_valido(linha.get(coluna_filial) if coluna_filial else None),
                "periodos": []
            }

        inicio = linha["Inicio"].date() if pd.notna(linha["Inicio"]) else None
        fim = linha["Fim"].date() if pd.notna(linha["Fim"]) else None
        bloqueio = linha["Data de Bloqueio"].date() if pd.notna(linha["Data de Bloqueio"]) else None
        if inicio and fim:
            grupos[chave]["periodos"].append({
                "inicio": inicio,
                "fim": fim,
                "bloqueio": bloqueio,
                "inicio_formatado": inicio.strftime("%d/%m/%Y"),
                "fim_formatado": fim.strftime("%d/%m/%Y"),
                "bloqueio_formatado": bloqueio.strftime("%d/%m/%Y") if bloqueio else "-"
            })

    colaboradores = []
    for chave, colaborador in grupos.items():
        perfil = substituicoes.get(chave, {})
        for campo in ["departamento", "cargo", "matricula", "filial"]:
            if perfil.get(campo):
                colaborador[campo] = perfil[campo]

        periodos = sorted(colaborador["periodos"], key=lambda item: item["inicio"])
        ativos = [item for item in periodos if item["inicio"] <= hoje <= item["fim"]]
        futuros = [item for item in periodos if item["inicio"] > hoje]
        passados = [item for item in periodos if item["fim"] < hoje]

        if ativos:
            periodo_atual = ativos[0]
            status, status_classe = "Em férias", "active"
        elif futuros:
            periodo_atual = futuros[0]
            status, status_classe = "Programada", "scheduled"
        elif passados:
            periodo_atual = passados[-1]
            status, status_classe = "Concluída", "completed"
        else:
            periodo_atual = None
            status, status_classe = "Sem período", "neutral"

        colaborador.update({
            "periodos": periodos,
            "periodo_atual": periodo_atual,
            "status": status,
            "status_classe": status_classe,
            "slug": quote(colaborador["nome"], safe=""),
            "iniciais": "".join(parte[0] for parte in colaborador["nome"].split()[:2]).upper(),
            "cor": sum(ord(letra) for letra in colaborador["nome"]) % 4
        })
        colaboradores.append(colaborador)

    return sorted(colaboradores, key=lambda item: normalizar_texto(item["nome"]))


@app.route("/colaboradores")
def colaboradores():
    dados = obter_colaboradores()
    agora = datetime.now()
    return render_template(
        "colaboradores.html",
        colaboradores=dados,
        total=len(dados),
        em_ferias=sum(1 for item in dados if item["status_classe"] == "active"),
        programadas=sum(1 for item in dados if item["status_classe"] == "scheduled"),
        departamentos=sorted({
            item["departamento"] for item in dados
            if item["departamento"] != "Não informado"
        }),
        total_departamentos=len({
            item["departamento"] for item in dados
            if item["departamento"] != "Não informado"
        }),
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y")
    )


@app.route("/colaboradores/<path:nome>")
def colaborador_detalhe(nome):
    colaborador = next(
        (item for item in obter_colaboradores() if normalizar_texto(item["nome"]) == normalizar_texto(nome)),
        None
    )
    if colaborador is None:
        return "Colaborador não encontrado.", 404

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT data_bloqueio, data_execucao
        FROM bloqueios
        WHERE lower(nome) = lower(?)
        ORDER BY datetime(data_execucao) DESC
    """, (colaborador["nome"],))
    bloqueios = cursor.fetchall()
    conn.close()

    historico_bloqueios = []
    acoes = []
    for data_bloqueio, data_execucao in bloqueios:
        try:
            bloqueio_dt = datetime.fromisoformat(str(data_bloqueio))
        except (TypeError, ValueError):
            bloqueio_dt = None
        try:
            execucao_dt = datetime.fromisoformat(str(data_execucao))
        except (TypeError, ValueError):
            execucao_dt = None
        historico_bloqueios.append({
            "data_bloqueio": bloqueio_dt.strftime("%d/%m/%Y") if bloqueio_dt else str(data_bloqueio or "-"),
            "data_execucao": execucao_dt.strftime("%d/%m/%Y %H:%M") if execucao_dt else str(data_execucao or "-")
        })
        if execucao_dt:
            acoes.append({
                "data": execucao_dt,
                "titulo": "Bloqueio executado",
                "descricao": f"Acesso bloqueado em {execucao_dt:%d/%m/%Y às %H:%M}",
                "tipo": "success"
            })

    for periodo in colaborador["periodos"]:
        acoes.extend([
            {
                "data": datetime.combine(periodo["inicio"], datetime.min.time()),
                "titulo": "Início das férias",
                "descricao": f"Período iniciado em {periodo['inicio_formatado']}",
                "tipo": "vacation"
            },
            {
                "data": datetime.combine(periodo["bloqueio"], datetime.min.time()) if periodo["bloqueio"] else datetime.combine(periodo["inicio"], datetime.min.time()),
                "titulo": "Bloqueio programado",
                "descricao": f"Bloqueio previsto para {periodo['bloqueio_formatado']}",
                "tipo": "scheduled"
            }
        ])

    anos = {}
    for periodo in colaborador["periodos"]:
        anos.setdefault(periodo["inicio"].year, []).append(periodo)

    agora = datetime.now()
    return render_template(
        "colaborador_detalhe.html",
        colaborador=colaborador,
        anos=sorted(anos.items(), reverse=True),
        historico_bloqueios=historico_bloqueios,
        acoes=sorted(acoes, key=lambda item: item["data"], reverse=True)[:6],
        editar=request.args.get("editar") == "1",
        salvo=request.args.get("salvo") == "1",
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y")
    )


@app.route("/colaboradores/<path:nome>/editar", methods=["POST"])
def colaborador_editar(nome):
    colaborador = next(
        (item for item in obter_colaboradores() if normalizar_texto(item["nome"]) == normalizar_texto(nome)),
        None
    )
    if colaborador is None:
        return "Colaborador não encontrado.", 404

    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO colaborador_perfis
        (nome, departamento, cargo, matricula, filial, atualizado_em)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(nome) DO UPDATE SET
            departamento=excluded.departamento,
            cargo=excluded.cargo,
            matricula=excluded.matricula,
            filial=excluded.filial,
            atualizado_em=excluded.atualizado_em
    """, (
        colaborador["nome"],
        request.form.get("departamento", "").strip(),
        request.form.get("cargo", "").strip(),
        request.form.get("matricula", "").strip(),
        request.form.get("filial", "").strip(),
        datetime.now().isoformat()
    ))
    conn.commit()
    conn.close()
    registrar_auditoria(
        "Editou colaborador",
        f"Perfil atualizado: {colaborador['nome']}"
    )
    return redirect(url_for("colaborador_detalhe", nome=colaborador["nome"], salvo="1"))

# ======================
# ▶ RODAR
# ======================
@app.route("/historico")
def historico():
    dados = obter_historico()
    agora = datetime.now()
    hoje_total = sum(
        1 for item in dados
        if item["execucao_dt"] and item["execucao_dt"].date() == agora.date()
    )
    mes_total = sum(
        1 for item in dados
        if item["execucao_dt"]
        and item["execucao_dt"].year == agora.year
        and item["execucao_dt"].month == agora.month
    )
    ultimo = dados[0] if dados else None

    return render_template(
        "historico.html",
        dados=dados,
        total=len(dados),
        hoje_total=hoje_total,
        mes_total=mes_total,
        ultimo=ultimo,
        timeline=dados[:5],
        busca=request.args.get("busca", ""),
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y"),
        data_hoje_dt=agora.date()
    )


def obter_historico():
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bloqueios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT,
            data_bloqueio TEXT,
            data_execucao TEXT
        )
    """)

    cursor.execute("""
        SELECT nome, data_bloqueio, data_execucao
        FROM bloqueios
        ORDER BY datetime(data_execucao) DESC
    """)
    registros = cursor.fetchall()
    conn.close()

    dados = []
    for nome, data_bloqueio, data_execucao in registros:
        try:
            bloqueio_dt = datetime.fromisoformat(str(data_bloqueio))
        except (TypeError, ValueError):
            bloqueio_dt = None

        try:
            execucao_dt = datetime.fromisoformat(str(data_execucao))
        except (TypeError, ValueError):
            execucao_dt = None

        dados.append({
            "nome": nome,
            "data_bloqueio": bloqueio_dt.strftime("%d/%m/%Y") if bloqueio_dt else str(data_bloqueio or "-"),
            "data_execucao": execucao_dt.strftime("%d/%m/%Y %H:%M") if execucao_dt else str(data_execucao or "-"),
            "execucao_hora": execucao_dt.strftime("%H:%M") if execucao_dt else "--:--",
            "execucao_data_curta": execucao_dt.strftime("%d/%m") if execucao_dt else "--/--",
            "execucao_dt": execucao_dt,
            "status": "Executado"
        })

    return dados


@app.route("/historico/exportar/excel")
def exportar_historico_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    dados = obter_historico()
    preferencias = carregar_configuracoes()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Histórico"
    sheet.append(["Colaborador", "Data do bloqueio", "Executado em", "Status"])

    for item in dados:
        sheet.append([
            item["nome"],
            item["data_bloqueio"],
            item["data_execucao"],
            item["status"]
        ])

    header_fill = PatternFill("solid", fgColor="082F5B")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")

    if preferencias.get("excel_congelar_cabecalho") == "1":
        sheet.freeze_panes = "A2"
    if preferencias.get("excel_autofiltro") == "1":
        sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 26
    if preferencias.get("excel_ajustar_colunas") == "1":
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 23
        sheet.column_dimensions["D"].width = 16

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    arquivo = BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)
    registrar_auditoria(
        "Exportou Excel",
        f"Histórico de bloqueios · {len(dados)} registro(s)"
    )

    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"historico_bloqueios_{datetime.now():%Y%m%d}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def gerar_pdf_historico(dados):
    preferencias = carregar_configuracoes()
    def pdf_texto(valor):
        texto = str(valor).encode("latin-1", "replace").decode("latin-1")
        return texto.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    linhas_por_pagina = 24
    paginas = [dados[i:i + linhas_por_pagina] for i in range(0, len(dados), linhas_por_pagina)] or [[]]
    objetos = {
        1: b"<< /Type /Catalog /Pages 2 0 R >>",
        3: b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>",
        4: b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold /Encoding /WinAnsiEncoding >>"
    }
    referencias_paginas = []
    proximo_objeto = 5

    for indice, pagina in enumerate(paginas, start=1):
        pagina_id = proximo_objeto
        conteudo_id = proximo_objeto + 1
        proximo_objeto += 2
        referencias_paginas.append(f"{pagina_id} 0 R")

        subtitulo = (
            f"Gerado em {datetime.now():%d/%m/%Y %H:%M}"
            if preferencias.get("pdf_data") == "1" else ""
        )
        comandos = [
            "0.032 0.184 0.357 rg 0 772 595 70 re f",
            "BT /F2 19 Tf 1 1 1 rg 42 806 Td (Historico de Bloqueios) Tj ET",
            f"BT /F1 9 Tf 0.82 0.9 1 rg 42 787 Td ({subtitulo}) Tj ET",
            "0.93 0.96 0.99 rg 42 736 511 25 re f",
            "BT /F2 9 Tf 0.18 0.25 0.34 rg 49 745 Td (COLABORADOR) Tj ET",
            "BT /F2 9 Tf 0.18 0.25 0.34 rg 250 745 Td (BLOQUEIO) Tj ET",
            "BT /F2 9 Tf 0.18 0.25 0.34 rg 350 745 Td (EXECUTADO EM) Tj ET",
            "BT /F2 9 Tf 0.18 0.25 0.34 rg 488 745 Td (STATUS) Tj ET"
        ]
        if preferencias.get("pdf_logo") == "1":
            comandos.append("BT /F2 9 Tf 0.45 0.72 1 rg 500 806 Td (FOKUS) Tj ET")

        y = 718
        for linha, item in enumerate(pagina):
            if linha % 2:
                comandos.append(f"0.97 0.98 0.99 rg 42 {y - 7} 511 25 re f")
            comandos.extend([
                f"BT /F1 9 Tf 0.16 0.22 0.3 rg 49 {y} Td ({pdf_texto(item['nome'])[:38]}) Tj ET",
                f"BT /F1 9 Tf 0.3 0.37 0.46 rg 250 {y} Td ({pdf_texto(item['data_bloqueio'])}) Tj ET",
                f"BT /F1 9 Tf 0.3 0.37 0.46 rg 350 {y} Td ({pdf_texto(item['data_execucao'])}) Tj ET",
                f"BT /F2 9 Tf 0.07 0.55 0.34 rg 488 {y} Td ({pdf_texto(item['status'])}) Tj ET"
            ])
            y -= 27

        if preferencias.get("pdf_rodape") == "1":
            comandos.append("BT /F1 8 Tf 0.45 0.5 0.58 rg 42 28 Td (Grupo Fokus Logistica - Sistema de Controle de Ferias) Tj ET")
        if preferencias.get("pdf_numero_pagina") == "1":
            comandos.append(f"BT /F1 8 Tf 0.45 0.5 0.58 rg 520 28 Td ({indice}/{len(paginas)}) Tj ET")
        stream = "\n".join(comandos).encode("latin-1")
        objetos[conteudo_id] = f"<< /Length {len(stream)} >>\nstream\n".encode("ascii") + stream + b"\nendstream"
        objetos[pagina_id] = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
            f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents {conteudo_id} 0 R >>"
        ).encode("ascii")

    objetos[2] = (
        f"<< /Type /Pages /Kids [{' '.join(referencias_paginas)}] /Count {len(referencias_paginas)} >>"
    ).encode("ascii")

    pdf = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0] * (max(objetos) + 1)
    for numero in range(1, max(objetos) + 1):
        offsets[numero] = len(pdf)
        pdf.extend(f"{numero} 0 obj\n".encode("ascii"))
        pdf.extend(objetos[numero])
        pdf.extend(b"\nendobj\n")

    xref = len(pdf)
    pdf.extend(f"xref\n0 {len(offsets)}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer\n<< /Size {len(offsets)} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode("ascii")
    )
    return bytes(pdf)


@app.route("/historico/exportar/pdf")
def exportar_historico_pdf():
    registrar_auditoria(
        "Gerou PDF",
        f"Histórico de bloqueios · {len(obter_historico())} registro(s)"
    )
    arquivo = BytesIO(gerar_pdf_historico(obter_historico()))
    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"historico_bloqueios_{datetime.now():%Y%m%d}.pdf",
        mimetype="application/pdf"
    )


@app.route("/relatorios")
def relatorios():
    dados = obter_dados_relatorios()
    return render_template(
        "relatorios.html",
        **dados,
        data_atual=dados["agora"].strftime("%d/%m/%Y %H:%M"),
        data_hoje=dados["agora"].strftime("%d/%m/%Y")
    )


def linhas_exportacao_relatorios(dados):
    return [
        [
            item["nome"],
            item["departamento"],
            item["inicio"].strftime("%d/%m/%Y"),
            item["fim"].strftime("%d/%m/%Y"),
            item["bloqueio"].strftime("%d/%m/%Y") if item["bloqueio"] else "-",
            item["dias"],
            item["status"]
        ]
        for item in dados["registros"]
    ]


@app.route("/relatorios/exportar/csv")
def exportar_relatorios_csv():
    dados = obter_dados_relatorios()
    texto = StringIO()
    texto.write("\ufeff")
    writer = csv.writer(texto, delimiter=";")
    writer.writerow([
        "Colaborador", "Departamento", "Início", "Fim",
        "Bloqueio", "Dias", "Status"
    ])
    writer.writerows(linhas_exportacao_relatorios(dados))
    arquivo = BytesIO(texto.getvalue().encode("utf-8"))
    registrar_auditoria(
        "Exportou CSV",
        f"Relatório gerencial · {dados['periodo_nome']} · {len(dados['registros'])} registro(s)"
    )
    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"relatorio_ferias_{dados['periodo'].replace('-', '_')}.csv",
        mimetype="text/csv; charset=utf-8"
    )


@app.route("/relatorios/exportar/excel")
def exportar_relatorios_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    dados = obter_dados_relatorios()
    preferencias = carregar_configuracoes()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Relatório de férias"
    sheet.append([
        "Colaborador", "Departamento", "Início", "Fim",
        "Bloqueio", "Dias", "Status"
    ])
    for linha in linhas_exportacao_relatorios(dados):
        sheet.append(linha)

    header_fill = PatternFill("solid", fgColor="082F5B")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    if preferencias.get("excel_congelar_cabecalho") == "1":
        sheet.freeze_panes = "A2"
    if preferencias.get("excel_autofiltro") == "1":
        sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 26
    if preferencias.get("excel_ajustar_colunas") == "1":
        for coluna, largura in {
            "A": 34, "B": 24, "C": 15, "D": 15,
            "E": 15, "F": 10, "G": 17
        }.items():
            sheet.column_dimensions[coluna].width = largura

    arquivo = BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)
    registrar_auditoria(
        "Exportou Excel",
        f"Relatório gerencial · {dados['periodo_nome']} · {len(dados['registros'])} registro(s)"
    )
    return send_file(
        arquivo,
        as_attachment=True,
        download_name=f"relatorio_ferias_{dados['periodo'].replace('-', '_')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def gerar_pdf_relatorios(dados):
    preferencias = carregar_configuracoes()
    def texto_pdf(valor):
        texto = str(valor).encode("latin-1", "replace").decode("latin-1")
        return texto.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    linhas = linhas_exportacao_relatorios(dados)
    paginas = [linhas[i:i + 24] for i in range(0, len(linhas), 24)] or [[]]
    objetos = {
        1: b"<< /Type /Catalog /Pages 2 0 R >>",
        3: b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>",
        4: b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold /Encoding /WinAnsiEncoding >>"
    }
    referencias = []
    proximo = 5

    for indice, pagina in enumerate(paginas, start=1):
        pagina_id, conteudo_id = proximo, proximo + 1
        proximo += 2
        referencias.append(f"{pagina_id} 0 R")
        comandos = [
            "0.032 0.184 0.357 rg 0 772 595 70 re f",
            "BT /F2 19 Tf 1 1 1 rg 42 806 Td (Relatorio Executivo de Ferias) Tj ET",
            f"BT /F1 9 Tf 0.82 0.9 1 rg 42 787 Td ({texto_pdf(dados['periodo_nome'])}) Tj ET",
            "0.93 0.96 0.99 rg 35 736 525 25 re f",
            "BT /F2 8 Tf 0.18 0.25 0.34 rg 41 745 Td (COLABORADOR) Tj ET",
            "BT /F2 8 Tf 0.18 0.25 0.34 rg 215 745 Td (DEPARTAMENTO) Tj ET",
            "BT /F2 8 Tf 0.18 0.25 0.34 rg 335 745 Td (INICIO) Tj ET",
            "BT /F2 8 Tf 0.18 0.25 0.34 rg 404 745 Td (FIM) Tj ET",
            "BT /F2 8 Tf 0.18 0.25 0.34 rg 471 745 Td (STATUS) Tj ET"
        ]
        if preferencias.get("pdf_logo") == "1":
            comandos.append("BT /F2 9 Tf 0.45 0.72 1 rg 500 806 Td (FOKUS) Tj ET")
        if preferencias.get("pdf_data") == "1":
            comandos.append(
                f"BT /F1 7 Tf 0.55 0.62 0.7 rg 430 787 Td (Gerado em {datetime.now():%d/%m/%Y}) Tj ET"
            )
        y = 718
        for linha_indice, linha in enumerate(pagina):
            if linha_indice % 2:
                comandos.append(f"0.97 0.98 0.99 rg 35 {y - 7} 525 25 re f")
            comandos.extend([
                f"BT /F1 8 Tf 0.16 0.22 0.3 rg 41 {y} Td ({texto_pdf(linha[0])[:38]}) Tj ET",
                f"BT /F1 8 Tf 0.3 0.37 0.46 rg 215 {y} Td ({texto_pdf(linha[1])[:25]}) Tj ET",
                f"BT /F1 8 Tf 0.3 0.37 0.46 rg 335 {y} Td ({texto_pdf(linha[2])}) Tj ET",
                f"BT /F1 8 Tf 0.3 0.37 0.46 rg 404 {y} Td ({texto_pdf(linha[3])}) Tj ET",
                f"BT /F2 8 Tf 0.07 0.45 0.34 rg 471 {y} Td ({texto_pdf(linha[6])[:18]}) Tj ET"
            ])
            y -= 27
        if preferencias.get("pdf_rodape") == "1":
            comandos.append("BT /F1 8 Tf 0.45 0.5 0.58 rg 35 28 Td (Grupo Fokus Logistica - Sistema de Controle de Ferias) Tj ET")
        if preferencias.get("pdf_numero_pagina") == "1":
            comandos.append(f"BT /F1 8 Tf 0.45 0.5 0.58 rg 520 28 Td ({indice}/{len(paginas)}) Tj ET")
        stream = "\n".join(comandos).encode("latin-1")
        objetos[conteudo_id] = (
            f"<< /Length {len(stream)} >>\nstream\n".encode("ascii")
            + stream + b"\nendstream"
        )
        objetos[pagina_id] = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
            f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> "
            f"/Contents {conteudo_id} 0 R >>"
        ).encode("ascii")

    objetos[2] = (
        f"<< /Type /Pages /Kids [{' '.join(referencias)}] /Count {len(referencias)} >>"
    ).encode("ascii")
    pdf = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0] * (max(objetos) + 1)
    for numero in range(1, max(objetos) + 1):
        offsets[numero] = len(pdf)
        pdf.extend(f"{numero} 0 obj\n".encode("ascii"))
        pdf.extend(objetos[numero])
        pdf.extend(b"\nendobj\n")
    xref = len(pdf)
    pdf.extend(f"xref\n0 {len(offsets)}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer\n<< /Size {len(offsets)} /Root 1 0 R >>\n"
        f"startxref\n{xref}\n%%EOF".encode("ascii")
    )
    return bytes(pdf)


@app.route("/relatorios/exportar/pdf")
def exportar_relatorios_pdf():
    dados = obter_dados_relatorios()
    registrar_auditoria(
        "Gerou PDF",
        f"Relatório gerencial · {dados['periodo_nome']} · {len(dados['registros'])} registro(s)"
    )
    return send_file(
        BytesIO(gerar_pdf_relatorios(dados)),
        as_attachment=True,
        download_name=f"relatorio_ferias_{dados['periodo'].replace('-', '_')}.pdf",
        mimetype="application/pdf"
    )


@app.route("/pesquisa")
def pesquisa_global():
    termo = request.args.get("q", "").strip()
    resultados = {
        "colaboradores": [],
        "calendario": [],
        "historico": [],
        "relatorios": []
    }
    if termo:
        termo_normalizado = normalizar_texto(termo)
        colaboradores = [
            item for item in obter_colaboradores()
            if termo_normalizado in normalizar_texto(item["nome"])
            or termo_normalizado in normalizar_texto(item["departamento"])
            or termo_normalizado in normalizar_texto(item["cargo"])
        ]
        for item in colaboradores[:20]:
            resultados["colaboradores"].append({
                "titulo": item["nome"],
                "descricao": f"{item['departamento']} · {item['status']}",
                "href": f"/colaboradores/{item['slug']}"
            })
            if item["periodo_atual"]:
                resultados["calendario"].append({
                    "titulo": item["nome"],
                    "descricao": (
                        f"{item['periodo_atual']['inicio_formatado']} a "
                        f"{item['periodo_atual']['fim_formatado']}"
                    ),
                    "href": f"/calendario?busca={quote(item['nome'])}"
                })

        for item in obter_historico():
            if termo_normalizado in normalizar_texto(item["nome"]):
                resultados["historico"].append({
                    "titulo": item["nome"],
                    "descricao": f"{item['status']} em {item['data_execucao']}",
                    "href": f"/historico?busca={quote(item['nome'])}"
                })
        resultados["historico"] = resultados["historico"][:20]

        if colaboradores:
            departamentos_encontrados = sorted({
                item["departamento"] for item in colaboradores
            })
            resultados["relatorios"].append({
                "titulo": f"Relatório filtrado por “{termo}”",
                "descricao": (
                    f"{len(colaboradores)} colaborador(es) · "
                    f"{', '.join(departamentos_encontrados[:3])}"
                ),
                "href": f"/relatorios?q={quote(termo)}"
            })
        inicializar_tabelas_sistema()
        conn = sqlite3.connect(DATABASE_PATH)
        logs = conn.execute(
            """
            SELECT acao, detalhe, usuario, data_hora
            FROM auditoria
            WHERE lower(acao) LIKE ? OR lower(detalhe) LIKE ?
               OR lower(usuario) LIKE ?
            ORDER BY datetime(data_hora) DESC
            LIMIT 12
            """,
            tuple([f"%{termo.lower()}%"] * 3)
        ).fetchall()
        conn.close()
        for acao, detalhe, usuario, data_hora in logs:
            try:
                momento = datetime.fromisoformat(data_hora)
                quando = momento.strftime("%d/%m às %H:%M")
            except (TypeError, ValueError):
                quando = "Data não informada"
            resultados["historico"].append({
                "titulo": acao,
                "descricao": f"{usuario} · {quando} · {detalhe or 'Sem detalhes'}",
                "href": f"/auditoria?q={quote(termo)}"
            })
        resultados["historico"] = resultados["historico"][:20]
        registrar_auditoria(
            "Realizou pesquisa global",
            f"Termo: {termo}"
        )

    total_resultados = sum(len(itens) for itens in resultados.values())
    agora = datetime.now()
    return render_template(
        "pesquisa.html",
        termo=termo,
        resultados=resultados,
        total_resultados=total_resultados,
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y")
    )


@app.route("/auditoria")
def auditoria():
    inicializar_tabelas_sistema()
    pesquisa = request.args.get("q", "").strip()
    acao_filtro = request.args.get("acao", "").strip()
    usuario_filtro = request.args.get("usuario", "").strip()
    data_inicio = request.args.get("data_inicio", "").strip()
    data_fim = request.args.get("data_fim", "").strip()
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    consulta = """
        SELECT id, data_hora, acao, detalhe, usuario, ip, resultado
        FROM auditoria WHERE 1=1
    """
    parametros = []
    if pesquisa:
        consulta += " AND (acao LIKE ? OR detalhe LIKE ? OR usuario LIKE ? OR ip LIKE ?)"
        termo = f"%{pesquisa}%"
        parametros.extend([termo, termo, termo, termo])
    if acao_filtro:
        consulta += " AND acao = ?"
        parametros.append(acao_filtro)
    if usuario_filtro:
        consulta += " AND usuario = ?"
        parametros.append(usuario_filtro)
    if data_inicio:
        consulta += " AND substr(data_hora, 1, 10) >= ?"
        parametros.append(data_inicio)
    if data_fim:
        consulta += " AND substr(data_hora, 1, 10) <= ?"
        parametros.append(data_fim)
    consulta += " ORDER BY datetime(data_hora) DESC LIMIT 300"
    registros = conn.execute(consulta, parametros).fetchall()
    acoes = [
        item[0] for item in conn.execute(
            "SELECT DISTINCT acao FROM auditoria ORDER BY acao"
        ).fetchall()
    ]
    usuarios = [
        item[0] for item in conn.execute(
            "SELECT DISTINCT usuario FROM auditoria ORDER BY usuario"
        ).fetchall()
    ]
    total = conn.execute("SELECT COUNT(*) FROM auditoria").fetchone()[0]
    hoje_iso = datetime.now().strftime("%Y-%m-%d")
    hoje_total = conn.execute(
        "SELECT COUNT(*) FROM auditoria WHERE substr(data_hora, 1, 10) = ?",
        (hoje_iso,)
    ).fetchone()[0]
    conn.close()

    dados = []
    for item in registros:
        try:
            momento = datetime.fromisoformat(item["data_hora"])
        except (TypeError, ValueError):
            momento = datetime.now()
        dados.append({
            "id": item["id"],
            "data": momento.strftime("%d/%m/%Y"),
            "hora": momento.strftime("%H:%M"),
            "acao": item["acao"],
            "detalhe": item["detalhe"] or "Sem detalhes adicionais",
            "usuario": item["usuario"],
            "ip": item["ip"] or "Não registrado",
            "resultado": item["resultado"] or "Sucesso"
        })

    agora = datetime.now()
    return render_template(
        "auditoria.html",
        registros=dados,
        acoes=acoes,
        usuarios=usuarios,
        pesquisa=pesquisa,
        acao_filtro=acao_filtro,
        usuario_filtro=usuario_filtro,
        data_inicio=data_inicio,
        data_fim=data_fim,
        total=total,
        hoje_total=hoje_total,
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y")
    )


@app.route("/configuracoes", methods=["GET", "POST"])
def configuracoes():
    erro = None
    if request.method == "POST":
        hora_rotina = request.form.get("hora_rotina", "08:00").strip()
        dias_antes = request.form.get("dias_antes_bloqueio", "3").strip()
        pasta_padrao = request.form.get("pasta_padrao", "").strip()
        manter_backups = request.form.get("manter_backups", "10").strip()
        try:
            datetime.strptime(hora_rotina, "%H:%M")
            dias_numero = int(dias_antes)
            manter_numero = int(manter_backups)
            if not 0 <= dias_numero <= 30:
                raise ValueError
            if not pasta_padrao or not 1 <= manter_numero <= 100:
                raise ValueError
        except ValueError:
            erro = "Revise os valores. A antecedência deve estar entre 0 e 30 dias."
        else:
            caixas = {
                "notificacoes_ativas", "som_notificacoes",
                "substituir_planilha", "validar_planilha",
                "mostrar_resumo_importacao", "backup_antes_importacao",
                "mostrar_popup", "notificar_importacao",
                "pdf_logo", "pdf_rodape", "pdf_numero_pagina", "pdf_data",
                "excel_autofiltro", "excel_ajustar_colunas",
                "excel_congelar_cabecalho",
                "limpeza_backups"
            }
            valores = {
                "nome_empresa": request.form.get(
                    "nome_empresa", "Grupo Fokus Logística"
                ).strip(),
                "nome_sistema": request.form.get(
                    "nome_sistema", "Fokus Férias"
                ).strip(),
                "versao": request.form.get("versao", "2.0").strip(),
                "banco": request.form.get("banco", "SQLite").strip(),
                "ambiente": request.form.get("ambiente", "Produção").strip(),
                "hora_rotina": hora_rotina,
                "dias_antes_bloqueio": str(dias_numero),
                "pasta_padrao": pasta_padrao,
                "tema": request.form.get("tema", "claro"),
                "cor_principal": request.form.get(
                    "cor_principal", "azul-fokus"
                ),
                "tamanho_fonte": request.form.get("tamanho_fonte", "normal"),
                "manter_backups": str(manter_numero),
            }
            valores.update({
                chave: "1" if request.form.get(chave) == "1" else "0"
                for chave in caixas
            })
            inicializar_tabelas_sistema()
            conn = sqlite3.connect(DATABASE_PATH)
            anteriores = dict(
                conn.execute("SELECT chave, valor FROM configuracoes").fetchall()
            )
            agora_iso = datetime.now().isoformat()
            for chave, valor in valores.items():
                conn.execute(
                    """
                    INSERT INTO configuracoes (chave, valor, atualizado_em)
                    VALUES (?, ?, ?)
                    ON CONFLICT(chave) DO UPDATE SET
                        valor=excluded.valor,
                        atualizado_em=excluded.atualizado_em
                    """,
                    (chave, valor, agora_iso)
                )
            conn.commit()
            conn.close()
            alteracoes = [
                chave for chave, valor in valores.items()
                if anteriores.get(chave) != valor
            ]
            nomes_campos = {
                "hora_rotina": "horário",
                "pasta_padrao": "pasta de importação",
                "dias_antes_bloqueio": "antecedência das notificações",
                "tema": "tema",
                "cor_principal": "cor principal",
                "tamanho_fonte": "tamanho da interface"
            }
            if alteracoes:
                resumo = ", ".join(
                    nomes_campos.get(chave, chave.replace("_", " "))
                    for chave in alteracoes[:5]
                )
                registrar_auditoria(
                    "Alterou configurações",
                    f"Campos alterados: {resumo}"
                )
            return redirect(url_for("configuracoes", salvo="1"))

    agora = datetime.now()
    backups = listar_backups()
    ultimo_backup = None
    if backups:
        momento = datetime.fromtimestamp(os.path.getmtime(backups[0]))
        ultimo_backup = {
            "data": momento.strftime("%d/%m/%Y"),
            "hora": momento.strftime("%H:%M"),
            "nome": os.path.basename(backups[0])
        }
    inicializar_tabelas_sistema()
    conn = sqlite3.connect(DATABASE_PATH)
    historico_configuracoes = conn.execute(
        """
        SELECT data_hora, detalhe, usuario
        FROM auditoria
        WHERE acao = 'Alterou configurações'
        ORDER BY datetime(data_hora) DESC
        LIMIT 6
        """
    ).fetchall()
    conn.close()
    historico = []
    for data_hora, detalhe, usuario in historico_configuracoes:
        try:
            momento = datetime.fromisoformat(data_hora)
        except (TypeError, ValueError):
            momento = agora
        historico.append({
            "data": momento.strftime("%d/%m"),
            "hora": momento.strftime("%H:%M"),
            "detalhe": detalhe,
            "usuario": usuario
        })
    return render_template(
        "configuracoes.html",
        configuracoes=carregar_configuracoes(),
        ultimo_backup=ultimo_backup,
        historico_configuracoes=historico,
        saude_sistema=obter_saude_sistema(),
        salvo=request.args.get("salvo") == "1",
        restaurado=request.args.get("restaurado") == "1",
        padrao_restaurado=request.args.get("padrao") == "1",
        teste_notificacao=request.args.get("teste_notificacao") == "1",
        backups_limpos=request.args.get("backups_limpos"),
        historico_limpo=request.args.get("historico_limpo"),
        erro_backup=request.args.get("erro_backup"),
        erro=erro,
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y")
    )


@app.route("/backup/gerar", methods=["POST"])
def gerar_backup():
    caminho_backup = criar_backup("manual")
    nome_arquivo = os.path.basename(caminho_backup)
    return send_file(
        caminho_backup,
        as_attachment=True,
        download_name=nome_arquivo,
        mimetype="application/zip"
    )


@app.route("/backup/limpar", methods=["POST"])
def limpar_backups():
    removidos = limpar_backups_antigos()
    registrar_auditoria(
        "Limpou backups antigos",
        f"{removidos} arquivo(s) removido(s)"
    )
    return redirect(url_for("configuracoes", backups_limpos=str(removidos)))


@app.route("/configuracoes/restaurar-padrao", methods=["POST"])
def restaurar_configuracoes_padrao():
    inicializar_tabelas_sistema()
    conn = sqlite3.connect(DATABASE_PATH)
    agora_iso = datetime.now().isoformat()
    for chave, valor in CONFIGURACOES_PADRAO.items():
        conn.execute(
            """
            INSERT INTO configuracoes (chave, valor, atualizado_em)
            VALUES (?, ?, ?)
            ON CONFLICT(chave) DO UPDATE SET
                valor=excluded.valor, atualizado_em=excluded.atualizado_em
            """,
            (chave, valor, agora_iso)
        )
    conn.commit()
    conn.close()
    registrar_auditoria(
        "Alterou configurações",
        "Restaurou todas as configurações padrão"
    )
    return redirect(url_for("configuracoes", padrao="1"))


@app.route("/configuracoes/testar-notificacao", methods=["POST"])
def testar_notificacao():
    registrar_auditoria(
        "Testou notificações",
        "Central de notificações respondeu corretamente"
    )
    return redirect(url_for("configuracoes", teste_notificacao="1"))


@app.route("/configuracoes/limpar-historico", methods=["POST"])
def limpar_historico_configuracoes():
    try:
        dias = int(request.form.get("dias", "90"))
    except ValueError:
        dias = 90
    if dias not in {30, 90, 180}:
        dias = 90
    limite = (datetime.now() - timedelta(days=dias)).isoformat()
    inicializar_tabelas_sistema()
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.execute(
        """
        DELETE FROM auditoria
        WHERE acao = 'Alterou configurações' AND data_hora < ?
        """,
        (limite,)
    )
    removidos = cursor.rowcount
    conn.commit()
    conn.close()
    registrar_auditoria(
        "Limpou histórico de configurações",
        f"Período: {dias} dias · {removidos} registro(s) removido(s)"
    )
    return redirect(url_for(
        "configuracoes", historico_limpo=str(removidos)
    ))


@app.route("/backup/restaurar", methods=["POST"])
def restaurar_backup():
    arquivo = request.files.get("backup")
    if not arquivo or not arquivo.filename.lower().endswith((".zip", ".db")):
        return redirect(url_for("configuracoes", erro_backup="arquivo"))
    try:
        criar_backup("pré-restauração")
        banco_temporario = DATABASE_PATH + ".restauracao"
        if arquivo.filename.lower().endswith(".db"):
            with open(banco_temporario, "wb") as destino:
                destino.write(arquivo.read())
        else:
            conteudo = BytesIO(arquivo.read())
            with zipfile.ZipFile(conteudo) as pacote:
                if "database/ferias.db" not in pacote.namelist():
                    raise ValueError
                with open(banco_temporario, "wb") as destino:
                    destino.write(pacote.read("database/ferias.db"))
        teste_conn = sqlite3.connect(banco_temporario)
        integridade = teste_conn.execute("PRAGMA integrity_check").fetchone()[0]
        teste_conn.close()
        if integridade != "ok":
            raise ValueError
        os.replace(banco_temporario, DATABASE_PATH)
        inicializar_tabelas_sistema()
        registrar_auditoria(
            "Restaurou backup",
            f"Arquivo: {os.path.basename(arquivo.filename)}"
        )
    except (zipfile.BadZipFile, KeyError, ValueError, OSError, sqlite3.Error):
        if os.path.exists(DATABASE_PATH + ".restauracao"):
            os.remove(DATABASE_PATH + ".restauracao")
        registrar_auditoria(
            "Restaurou backup",
            f"Falha no arquivo: {os.path.basename(arquivo.filename)}",
            resultado="Falha"
        )
        return redirect(url_for("configuracoes", erro_backup="invalido"))
    return redirect(url_for("configuracoes", restaurado="1"))


@app.route("/sobre")
def sobre():
    agora = datetime.now()
    return render_template(
        "sobre.html",
        configuracoes=carregar_configuracoes(),
        data_atual=agora.strftime("%d/%m/%Y %H:%M"),
        data_hoje=agora.strftime("%d/%m/%Y")
    )


@app.route("/api/versao-dados")
def versao_dados():
    caminho = planilha_mais_recente()
    if not caminho or not os.path.exists(caminho):
        return {"versao": "sem-planilha"}
    estado = os.stat(caminho)
    return {
        "versao": f"{estado.st_mtime_ns}-{estado.st_size}",
        "atualizado_em": datetime.fromtimestamp(
            estado.st_mtime
        ).isoformat(timespec="seconds")
    }


@app.route("/manutencao")
def manutencao():
    return render_template(
        "estado_sistema.html",
        codigo="FOKUS FÉRIAS",
        titulo="Estamos realizando uma atualização",
        mensagem="Voltaremos em instantes. Esta página será atualizada automaticamente.",
        acao_href=None,
        acao_icone=None,
        acao_texto=None,
        atualizar=True
    ), 503


@app.before_request
def verificar_modo_manutencao():
    if (
        os.environ.get("FOKUS_MODO_MANUTENCAO") == "1"
        and request.endpoint not in {"manutencao", "static"}
    ):
        return manutencao()


@app.errorhandler(404)
def pagina_nao_encontrada(_erro):
    return render_template(
        "estado_sistema.html",
        codigo="404",
        titulo="Página não encontrada",
        mensagem="O endereço informado não existe ou foi movido.",
        acao_href=url_for("dashboard"),
        acao_icone="layout-dashboard",
        acao_texto="Voltar ao Dashboard",
        atualizar=False
    ), 404


if __name__ == "__main__":
    app.run(debug=True)


# rota para servir imagens da pasta Imagens (permite acessar a logo existente)
@app.route('/imagens/<path:filename>')
def imagens(filename):
    return send_from_directory('Imagens', filename)


@app.route('/uploads/<path:filename>')
def uploads_files(filename):
    return send_from_directory(obter_pasta_planilhas(), filename)
